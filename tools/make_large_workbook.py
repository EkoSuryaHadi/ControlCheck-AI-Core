#!/usr/bin/env python3
"""Generate a large valid EPC workbook (>4 MiB) by replicating the realistic
dataset rows many times — used to prove the async worker path handles
workbooks far above the old 4 MiB sync limit."""
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

SRC = Path("/home/ubuntu/ControlCheck-AI-Core/data/ControlCheck_AI_Realistic_EPC_Dummy_Dataset_v1.0.xlsx")
OUT = Path("/home/ubuntu/ControlCheck-AI-Core/var/large_test_workbook.xlsx")


def main() -> Path:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(SRC, read_only=False, data_only=True)
    target = 8 * 1024 * 1024  # comfortably >4 MiB
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        header, data = rows[0], rows[1:]
        if not data:
            continue
        # Keep meta sheets at original scale so ingestion validation passes;
        # inflate only the transactional sheets that feed the audit engine.
        if ws.title in ("README", "Project_Info", "Scenario_Notes", "WBS"):
            print(f"  sheet '{ws.title}': kept at {len(data)} rows (meta)")
            continue
        repeat = max(2, target // max(1, len(header) * 30 * len(data)))
        repeat = min(repeat, 2000)  # sanity cap
        new_rows = [header]
        for _ in range(repeat):
            new_rows.extend(data)
        ws.delete_rows(1, ws.max_row)
        for r in new_rows:
            ws.append(r)
        print(f"  sheet '{ws.title}': {len(data)} → {len(new_rows)-1} rows (x{repeat})")
    wb.save(OUT)
    size = OUT.stat().st_size
    print(f"Saved {OUT} ({size/1024/1024:.1f} MiB)")
    assert size > 4 * 1024 * 1024, "test file must exceed the 4 MiB sync limit"
    return OUT


if __name__ == "__main__":
    main()
