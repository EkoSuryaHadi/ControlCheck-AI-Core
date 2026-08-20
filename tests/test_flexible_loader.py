from io import BytesIO
from datetime import date
import openpyxl

from controlcheck.loader import load_workbook
from controlcheck.config import load_catalogue, ThresholdConfig
from controlcheck.engine import ControlEngine, RuleContext
from controlcheck.rules import ALL_RULES


def _create_custom_indonesian_workbook() -> BytesIO:
    wb = openpyxl.Workbook()
    # Sheet 1: Info Proyek
    ws_info = wb.active
    ws_info.title = "Info Proyek"
    ws_info.append(["Parameter", "Nilai"])
    ws_info.append(["Project ID", "PRJ-INDO-001"])
    ws_info.append(["Project Name", "Proyek Konstruksi Pabrik"])
    ws_info.append(["Data Date", "2026-08-15"])
    ws_info.append(["Dataset Version", "0.2"])

    # Sheet 2: Struktur WBS
    ws_wbs = wb.create_sheet("Struktur WBS")
    ws_wbs.append(["Kode WBS", "Nama WBS", "Induk WBS", "Disiplin", "Tingkat"])
    ws_wbs.append(["1.0", "Pekerjaan Sipil", None, "CIVIL", 1])
    ws_wbs.append(["1.1", "Pondasi Bangunan", "1.0", "CIVIL", 2])

    # Sheet 3: Anggaran RAB
    ws_bdg = wb.create_sheet("Anggaran RAB")
    ws_bdg.append(["Kode WBS", "Kode Akun", "Nilai RAB", "Status", "Tgl Berlaku"])
    ws_bdg.append(["1.1", "MAT-01", 1000000000, "APPROVED", "2026-01-01"])

    # Sheet 4: Realisasi Biaya
    ws_act = wb.create_sheet("Realisasi Biaya")
    ws_act.append(["No Transaksi", "Kode WBS", "Kode Akun", "Biaya Aktual", "Nama Vendor", "Tgl Transaksi"])
    ws_act.append(["TX-001", "1.1", "MAT-01", 1250000000, "PT Semen Nusantara", "2026-08-10"])

    # Sheet 5: Jadwal Proyek
    ws_sch = wb.create_sheet("Jadwal Proyek")
    ws_sch.append(["ID Aktivitas", "Kode WBS", "Nama Aktivitas", "Tgl Mulai Rencana", "Tgl Selesai Rencana", "Bobot Rencana %", "Bobot Realisasi %", "Total Float", "Kritis"])
    ws_sch.append(["ACT-001", "1.1", "Pengecoran Pondasi", "2026-01-10", "2026-08-01", 100, 50, 0, "Ya"])

    # Sheet 6: Kurva S Progres
    ws_prg = wb.create_sheet("Kurva S")
    ws_prg.append(["Kode WBS", "Periode", "Rencana %", "Realisasi %"])
    ws_prg.append(["1.1", "2026-08-15", 100, 50])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_flexible_loader_parses_indonesian_sheets():
    buf = _create_custom_indonesian_workbook()
    dataset = load_workbook(buf)

    assert dataset.project.project_id == "PRJ-INDO-001"
    assert len(dataset.wbs_nodes) == 2
    assert len(dataset.budgets) == 1
    assert len(dataset.actual_costs) == 1
    assert dataset.actual_costs[0].actual_amount == 1250000000
    assert len(dataset.schedule) == 1
    assert dataset.schedule[0].critical is True
    assert len(dataset.progress) == 1


def test_flexible_dataset_evaluates_rules_successfully():
    buf = _create_custom_indonesian_workbook()
    dataset = load_workbook(buf)

    catalogue = load_catalogue("data/controlcheck_rule_catalogue_v0.2.json")
    context = RuleContext(catalogue=catalogue, thresholds=ThresholdConfig())
    engine = ControlEngine(ALL_RULES)
    result = engine.run(dataset, context)

    # Actual cost (1.25B) exceeds budget (1.0B) => CST-001 should trigger
    cst_001_findings = [f for f in result.findings if f.rule_id == "CST-001"]
    assert len(cst_001_findings) >= 1
