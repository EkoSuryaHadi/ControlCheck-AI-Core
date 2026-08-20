from pathlib import Path

import openpyxl
import pytest

from controlcheck.loader import WorkbookSchemaError, load_workbook


def test_loader_reads_project_and_preserves_source_rows(sample_workbook: Path):
    dataset = load_workbook(sample_workbook)

    assert dataset.project.project_id == "PRJ-CCAI-001"
    assert dataset.data_date.isoformat() == "2026-08-15"
    planted = next(x for x in dataset.actual_costs if x.transaction_id == "ACT-9003")
    assert planted.wbs_code is None
    assert planted.source.sheet == "Actual_Cost"
    assert planted.source.row_number == 69


def test_loader_normalizes_dates_percentages_and_money(sample_workbook: Path):
    dataset = load_workbook(sample_workbook)

    schedule = next(x for x in dataset.schedule if x.activity_id == "A9990")
    progress = next(x for x in dataset.progress if x.progress_id == "PRG-50-4")
    transaction = next(x for x in dataset.actual_costs if x.transaction_id == "ACT-9006")
    assert schedule.actual_finish.isoformat() == "2026-06-10"
    assert progress.actual_progress == pytest.approx(1.08)
    assert int(transaction.actual_amount) == 4_600_000_000


def test_loader_rejects_missing_required_sheet(tmp_path: Path, sample_workbook: Path):
    book = openpyxl.load_workbook(sample_workbook)
    del book["Budget"]
    invalid = tmp_path / "missing-budget.xlsx"
    book.save(invalid)

    with pytest.raises(WorkbookSchemaError, match="Budget") as exc:
        load_workbook(invalid, strict=True)
    assert exc.value.code == "missing_sheets"


def test_loader_rejects_missing_required_column(tmp_path: Path, sample_workbook: Path):
    book = openpyxl.load_workbook(sample_workbook)
    book["Actual_Cost"]["I3"] = "wrong_amount_column"
    invalid = tmp_path / "missing-column.xlsx"
    book.save(invalid)

    with pytest.raises(WorkbookSchemaError, match="actual_amount") as exc:
        load_workbook(invalid, strict=True)
    assert exc.value.code == "missing_columns"

