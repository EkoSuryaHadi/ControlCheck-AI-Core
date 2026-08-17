from decimal import Decimal
from pathlib import Path

import openpyxl

from controlcheck.loader import load_workbook


NUMERIC_RULES = {
    "DQ-003",
    "CST-001", "CST-002", "CST-003", "CST-004", "CST-005", "CST-006",
    "SCH-001", "SCH-002", "SCH-003", "SCH-004", "SCH-005",
    "PRG-001", "PRG-002", "PRG-003",
    "XDOM-001",
}


def _paths(project_root: Path) -> tuple[Path, Path]:
    return (
        project_root / "data" / "ControlCheck_AI_Golden_Positive_Dataset_v0.2.xlsx",
        project_root / "data" / "ControlCheck_AI_Boundary_Negative_Dataset_v0.2.xlsx",
    )


def _validation_rows(path: Path) -> list[dict]:
    book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = book["Validation_Cases"]
    values = list(sheet.iter_rows(values_only=True))
    header_index = next(
        index for index, row in enumerate(values)
        if row and row[0] == "case_id"
    )
    headers = list(values[header_index])
    rows = [
        dict(zip(headers, row))
        for row in values[header_index + 1:]
        if row and row[0]
    ]
    book.close()
    return rows


def test_v02_workbooks_load_and_declare_dataset_version(project_root: Path):
    for path in _paths(project_root):
        dataset = load_workbook(path)
        assert dataset.dataset_version == "0.2"


def test_golden_planted_values_satisfy_their_formulas(project_root: Path):
    golden, _ = _paths(project_root)
    dataset = load_workbook(golden)
    budgets = {item.wbs_code: item.budget_amount for item in dataset.budgets}
    actuals = {
        wbs: sum(
            (item.actual_amount for item in dataset.actual_costs if item.wbs_code == wbs),
            Decimal("0"),
        )
        for wbs in {"3.1", "3.2"}
    }

    assert budgets["3.1"] == Decimal("8000000000")
    assert actuals["3.1"] > budgets["3.1"]
    assert budgets["3.2"] == Decimal("16000000000")
    assert actuals["3.2"] > budgets["3.2"]
    commitment = next(item for item in dataset.commitments if item.commitment_id == "COM-005")
    latest = next(item for item in dataset.progress if item.progress_id == "PRG-31-4")
    assert commitment.invoiced_amount == Decimal("5000000000")
    assert latest.actual_progress == 0.60


def test_boundary_dataset_has_literal_below_equal_above_cases(project_root: Path):
    _, boundary = _paths(project_root)
    rows = _validation_rows(boundary)

    assert {row["rule_id"] for row in rows} == NUMERIC_RULES
    for rule_id in NUMERIC_RULES:
        types = {row["boundary_type"] for row in rows if row["rule_id"] == rule_id}
        assert {"below", "equal", "above"} <= types
    assert {
        row["boundary_type"] for row in rows if row["rule_id"] == "CST-005"
    } == {"below", "equal", "above", "approved_exception"}
    assert all(
        isinstance(row["input_value"], (int, float))
        and isinstance(row["threshold_value"], (int, float))
        for row in rows
    )


def test_boundary_expectations_respect_inclusive_and_strict_operators(project_root: Path):
    _, boundary = _paths(project_root)
    rows = {(row["rule_id"], row["boundary_type"]): row for row in _validation_rows(boundary)}

    assert rows[("CST-001", "equal")]["expected_trigger"] is False
    assert rows[("CST-003", "equal")]["expected_trigger"] is True
    assert rows[("SCH-005", "below")]["expected_trigger"] is True
    assert rows[("SCH-005", "equal")]["expected_trigger"] is False
    assert rows[("PRG-002", "above")]["expected_trigger"] is True

