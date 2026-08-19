from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path
from uuid import uuid4

import openpyxl
import pytest
from alembic import command
from sqlalchemy import select

from controlcheck.errors import ControlCheckApplicationError
from controlcheck.ingestion.mapper import DomainStatus
from controlcheck.ingestion.profile import load_mapping_profile, mapping_profile_sha256
from controlcheck.ingestion.service import SnapshotIngestionService
from controlcheck.loader import load_workbook
from controlcheck.persistence.database import create_session_factory
from controlcheck.persistence.dataset_loader import DatabaseDatasetLoader
from controlcheck.persistence.ingestion_repositories import SnapshotRepository
from controlcheck.persistence.models import (
    OrganizationRecord,
    ProjectRecord,
    RawRowRecord,
    SourceFileRecord,
)
from controlcheck.storage import LocalFileStorage


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture()
def session_factory(alembic_config, postgres_url):
    command.upgrade(alembic_config, "head")
    return create_session_factory(postgres_url)


@pytest.fixture(scope="session")
def golden_path(project_root: Path) -> Path:
    return project_root / "data" / "ControlCheck_AI_Golden_Positive_Dataset_v0.2.xlsx"


@pytest.fixture(scope="session")
def golden_bytes(golden_path: Path) -> bytes:
    return golden_path.read_bytes()


@pytest.fixture(scope="session")
def golden_dataset(golden_path: Path):
    return load_workbook(golden_path)


@pytest.fixture()
def golden_project(session_factory, golden_dataset) -> ProjectRecord:
    with session_factory() as session:
        organization = OrganizationRecord(
            name="Database dataset loader organization",
            slug=f"database-loader-{uuid4().hex}",
        )
        session.add(organization)
        session.flush()
        project = ProjectRecord(
            organization_id=organization.id,
            code=golden_dataset.project.project_id,
            # The mutable project master name must not become snapshot evidence.
            name="Golden project",
            currency="IDR",
        )
        session.add(project)
        session.commit()
        return project


@pytest.fixture()
def mapping_profile(project_root: Path):
    return load_mapping_profile(project_root / "data" / "controlcheck_mapping_profile_v0.1.json")


@pytest.fixture()
def storage(tmp_path: Path):
    return LocalFileStorage(tmp_path)


@pytest.fixture()
def snapshot_service(session_factory, storage, mapping_profile):
    return SnapshotIngestionService(session_factory, storage, mapping_profile)


@pytest.fixture()
def ingested_golden(snapshot_service, golden_project, golden_bytes):
    return snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )


@pytest.fixture()
def db_loader(session_factory):
    return DatabaseDatasetLoader(session_factory)


def _begin_snapshot(
    session_factory,
    storage,
    mapping_profile,
    golden_bytes,
    golden_project,
):
    with session_factory() as session:
        repository = SnapshotRepository(session)
        profile_record = repository.resolve_mapping_profile(
            mapping_profile,
            mapping_profile_sha256(mapping_profile),
        )
        stored = storage.put(
            golden_project.organization_id,
            golden_project.id,
            f"unfinished-{uuid4().hex}.xlsx",
            golden_bytes,
        )
        snapshot = repository.create_ingesting(
            organization_id=golden_project.organization_id,
            project_id=golden_project.id,
            filename="unfinished.xlsx",
            content_type=XLSX_MIME,
            stored=stored,
            mapping_profile_version_id=profile_record.id,
            dataset_version="0.2",
            data_date=date(2026, 8, 15),
            source_project_id=golden_project.code,
            source_project_name="EPC Gas Compression Facility Expansion",
            dedupe_key=None,
        )
        session.commit()
        return snapshot


def _assert_application_error(
    error: ControlCheckApplicationError,
    code: str,
    message: str,
    status_code: int,
) -> None:
    assert (error.code, error.message, error.status_code) == (code, message, status_code)


def test_db_loader_matches_legacy_loader_exactly(
    ingested_golden,
    golden_project,
    golden_dataset,
    db_loader,
):
    actual = db_loader.load(
        golden_project.organization_id,
        golden_project.id,
        ingested_golden.id,
    )

    assert actual.snapshot.model_dump(mode="json") == golden_dataset.model_dump(mode="json")


@pytest.mark.parametrize(
    "source_project_name",
    ["  EPC Gas Compression Facility Expansion  ", "P" * 300],
)
def test_db_loader_preserves_source_project_name_losslessly(
    source_project_name,
    golden_bytes,
    golden_project,
    snapshot_service,
    db_loader,
    tmp_path,
):
    workbook = openpyxl.load_workbook(BytesIO(golden_bytes))
    workbook["Project_Info"]["B3"] = source_project_name
    stream = BytesIO()
    workbook.save(stream)
    changed_bytes = stream.getvalue()
    changed_path = tmp_path / "project-name.xlsx"
    changed_path.write_bytes(changed_bytes)

    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        changed_path.name,
        XLSX_MIME,
        changed_bytes,
    )
    expected = load_workbook(changed_path)
    actual = db_loader.load(
        golden_project.organization_id,
        golden_project.id,
        snapshot.id,
    )

    assert actual.snapshot.project == expected.project


def test_db_loader_never_reads_source_file(
    ingested_golden,
    golden_project,
    db_loader,
    storage,
    session_factory,
):
    with session_factory() as session:
        source = session.scalar(
            select(SourceFileRecord).where(SourceFileRecord.id == ingested_golden.source_file_id)
        )
    assert source is not None
    storage.delete(source.storage_key)

    loaded = db_loader.load(
        golden_project.organization_id,
        golden_project.id,
        ingested_golden.id,
    )

    assert len(loaded.snapshot.actual_costs) == 73
    assert sum(
        len(records)
        for records in (
            loaded.snapshot.wbs_nodes,
            loaded.snapshot.budgets,
            loaded.snapshot.actual_costs,
            loaded.snapshot.commitments,
            loaded.snapshot.schedule,
            loaded.snapshot.progress,
        )
    ) == 149


def test_db_loader_returns_domain_statuses_and_raw_row_lineage(
    ingested_golden,
    golden_project,
    db_loader,
    session_factory,
):
    loaded = db_loader.load(
        golden_project.organization_id,
        golden_project.id,
        ingested_golden.id,
    )
    with session_factory() as session:
        rows = session.scalars(
            select(RawRowRecord).where(
                RawRowRecord.organization_id == golden_project.organization_id,
                RawRowRecord.project_id == golden_project.id,
                RawRowRecord.dataset_snapshot_id == ingested_golden.id,
            )
        ).all()

    expected_index = {
        (row.source_sheet, row.source_row_number): row.id
        for row in rows
    }
    assert loaded.domain_statuses == {
        "actual_cost": DomainStatus.valid,
        "budget": DomainStatus.valid,
        "commitments": DomainStatus.valid,
        "progress": DomainStatus.valid,
        "schedule": DomainStatus.valid,
        "wbs": DomainStatus.valid,
    }
    assert loaded.raw_row_index == expected_index
    assert len(loaded.raw_row_index) == 149
    assert all(isinstance(raw_row_id, int) for raw_row_id in loaded.raw_row_index.values())


def test_db_loader_preserves_rule_detectable_anomalies(
    ingested_golden,
    golden_project,
    db_loader,
):
    loaded = db_loader.load(
        golden_project.organization_id,
        golden_project.id,
        ingested_golden.id,
    )

    progress = next(row for row in loaded.snapshot.progress if row.progress_id == "PRG-50-4")
    contradictory = next(row for row in loaded.snapshot.schedule if row.activity_id == "A9990")
    assert progress.actual_progress == 1.08
    assert progress.variance == 0.20000000000000007
    assert contradictory.actual_start == date(2026, 6, 20)
    assert contradictory.actual_finish == date(2026, 6, 10)


@pytest.mark.parametrize("wrong_scope", ["organization", "project"])
def test_db_loader_snapshot_lookup_is_tenant_and_project_scoped(
    wrong_scope,
    ingested_golden,
    golden_project,
    db_loader,
):
    organization_id = (
        uuid4() if wrong_scope == "organization" else golden_project.organization_id
    )
    project_id = uuid4() if wrong_scope == "project" else golden_project.id

    with pytest.raises(ControlCheckApplicationError) as caught:
        db_loader.load(organization_id, project_id, ingested_golden.id)

    _assert_application_error(
        caught.value,
        "snapshot_not_found",
        "Dataset snapshot was not found for this project",
        404,
    )


def test_db_loader_rejects_ingesting_snapshot_with_stable_error(
    session_factory,
    storage,
    mapping_profile,
    golden_bytes,
    golden_project,
    db_loader,
):
    snapshot = _begin_snapshot(
        session_factory,
        storage,
        mapping_profile,
        golden_bytes,
        golden_project,
    )

    with pytest.raises(ControlCheckApplicationError) as caught:
        db_loader.load(golden_project.organization_id, golden_project.id, snapshot.id)

    _assert_application_error(
        caught.value,
        "snapshot_not_ready",
        "Dataset snapshot is still ingesting",
        409,
    )


def test_db_loader_rejects_failed_snapshot_with_stable_error(
    session_factory,
    storage,
    mapping_profile,
    golden_bytes,
    golden_project,
    db_loader,
):
    snapshot = _begin_snapshot(
        session_factory,
        storage,
        mapping_profile,
        golden_bytes,
        golden_project,
    )
    with session_factory() as session:
        SnapshotRepository(session).fail(
            golden_project.organization_id,
            golden_project.id,
            snapshot.id,
            "test_failure",
            "Test failure",
        )
        session.commit()

    with pytest.raises(ControlCheckApplicationError) as caught:
        db_loader.load(golden_project.organization_id, golden_project.id, snapshot.id)

    _assert_application_error(
        caught.value,
        "snapshot_failed",
        "Dataset snapshot ingestion failed",
        409,
    )
