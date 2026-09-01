from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from uuid import uuid4

import openpyxl
import pytest
from alembic import command
from sqlalchemy import and_, func, select

import controlcheck.ingestion.service as ingestion_service_module
from controlcheck.errors import ControlCheckApplicationError
from controlcheck.ingestion.extractor import extract_workbook
from controlcheck.ingestion.mapper import map_extracted_workbook
from controlcheck.ingestion.profile import load_mapping_profile, mapping_profile_sha256
from controlcheck.ingestion.service import SnapshotIngestionService
from controlcheck.persistence.database import create_session_factory
from controlcheck.persistence.ingestion_repositories import (
    SnapshotImmutableError,
    SnapshotRepository,
)
from controlcheck.persistence.models import (
    GovernedActualCostRecord,
    GovernedBudgetRecord,
    GovernedCommitmentRecord,
    GovernedDatasetDomainStatusRecord,
    GovernedDatasetSnapshotRecord,
    GovernedImportBatchRecord,
    GovernedMappingProfileVersionRecord,
    OrganizationRecord,
    GovernedProgressRecord,
    ProjectRecord,
    GovernedRawRowRecord,
    GovernedScheduleActivityRecord,
    SourceFileRecord,
    GovernedWBSNodeRecord,
)
from controlcheck.storage import LocalFileStorage


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CANONICAL_MODELS = (
    GovernedWBSNodeRecord,
    GovernedBudgetRecord,
    GovernedActualCostRecord,
    GovernedCommitmentRecord,
    GovernedScheduleActivityRecord,
    GovernedProgressRecord,
)


@pytest.fixture()
def session_factory(alembic_config, postgres_url):
    command.upgrade(alembic_config, "head")
    return create_session_factory(postgres_url)


@pytest.fixture()
def db_session(session_factory):
    with session_factory() as session:
        yield session


@pytest.fixture()
def golden_project(session_factory) -> ProjectRecord:
    with session_factory() as session:
        organization = OrganizationRecord(
            name="Snapshot ingestion organization",
            slug=f"snapshot-ingestion-{uuid4().hex}",
        )
        session.add(organization)
        session.flush()
        project = ProjectRecord(
            organization_id=organization.id,
            code="PRJ-CCAI-001",
            name="Golden project",
            currency="IDR",
        )
        session.add(project)
        session.commit()
        return project


@pytest.fixture(scope="session")
def golden_bytes(project_root: Path) -> bytes:
    return (project_root / "data" / "ControlCheck_AI_Golden_Positive_Dataset_v0.2.xlsx").read_bytes()


@pytest.fixture()
def mapping_profile(project_root: Path):
    return load_mapping_profile(project_root / "data" / "controlcheck_mapping_profile_v0.1.json")


@pytest.fixture()
def storage(tmp_path: Path):
    return LocalFileStorage(tmp_path)


@pytest.fixture()
def snapshot_service(session_factory, storage, mapping_profile):
    return SnapshotIngestionService(
        session_factory=session_factory,
        storage=storage,
        mapping_profile=mapping_profile,
    )


def count_rows(db_session, model, snapshot_id) -> int:
    return db_session.scalar(
        select(func.count()).select_from(model).where(model.dataset_snapshot_id == snapshot_id)
    )


def count_scoped(db_session, model, organization_id, project_id) -> int:
    return db_session.scalar(
        select(func.count()).select_from(model).where(
            model.organization_id == organization_id,
            model.project_id == project_id,
        )
    )


def stored_files(storage: LocalFileStorage) -> list[Path]:
    return sorted(path for path in storage.root.rglob("*") if path.is_file())


def mutate_workbook(data: bytes, mutation) -> bytes:
    book = openpyxl.load_workbook(BytesIO(data))
    try:
        mutation(book)
        output = BytesIO()
        book.save(output)
        return output.getvalue()
    finally:
        book.close()


def replace_project_id(book, value: str) -> None:
    sheet = book["Project_Info"]
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == "project_id":
            sheet.cell(row=row, column=2, value=value)
            return
    raise AssertionError("Golden workbook has no project_id metadata row")


def begin_ingesting_snapshot(
    session,
    storage,
    mapping_profile,
    golden_bytes,
    golden_project,
    dedupe_key,
):
    extracted = extract_workbook(golden_bytes, mapping_profile)
    mapped = map_extracted_workbook(extracted, mapping_profile)
    repository = SnapshotRepository(session)
    profile_record = repository.resolve_mapping_profile(
        mapping_profile,
        mapping_profile_sha256(mapping_profile),
    )
    stored = storage.put(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        golden_bytes,
    )
    snapshot = repository.create_ingesting(
        organization_id=golden_project.organization_id,
        project_id=golden_project.id,
        filename="golden.xlsx",
        content_type=XLSX_MIME,
        stored=stored,
        mapping_profile_version_id=profile_record.id,
        dataset_version="0.2",
        data_date=date(2026, 8, 15),
        source_project_id="PRJ-CCAI-001",
        source_project_name="EPC Gas Compression Facility Expansion",
        dedupe_key=dedupe_key,
    )
    return repository, snapshot, stored, extracted, mapped


def test_golden_ingestion_persists_raw_and_canonical_rows(
    snapshot_service, golden_bytes, golden_project, db_session
):
    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )

    assert snapshot.status == "validated"
    assert snapshot.row_count_raw == 149
    assert snapshot.row_count_canonical == 149
    assert count_rows(db_session, GovernedRawRowRecord, snapshot.id) == 149
    assert sum(count_rows(db_session, model, snapshot.id) for model in CANONICAL_MODELS) == 149


def test_duplicate_upload_returns_existing_snapshot(
    snapshot_service, golden_bytes, golden_project, storage
):
    first = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )
    second = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "renamed.xlsx",
        XLSX_MIME,
        golden_bytes,
    )

    assert first.outcome == "created"
    assert second.outcome == "deduplicated"
    assert second.snapshot.id == first.snapshot.id
    assert len(stored_files(storage)) == 1


def test_force_new_creates_distinct_snapshot(snapshot_service, golden_bytes, golden_project):
    first = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )
    forced = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
        force_new=True,
    )

    assert first.outcome == "created"
    assert forced.outcome == "created"
    assert forced.snapshot.id != first.snapshot.id
    assert forced.snapshot.dedupe_key is None


def test_golden_rows_keep_exact_lineage_and_rule_detectable_anomalies(
    snapshot_service, golden_bytes, golden_project, db_session, storage
):
    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )

    for model in CANONICAL_MODELS:
        linked = db_session.execute(
            select(model, GovernedRawRowRecord)
            .join(
                GovernedRawRowRecord,
                and_(
                    GovernedRawRowRecord.dataset_snapshot_id == model.dataset_snapshot_id,
                    GovernedRawRowRecord.id == model.raw_row_id,
                ),
            )
            .where(
                model.organization_id == golden_project.organization_id,
                model.project_id == golden_project.id,
                model.dataset_snapshot_id == snapshot.id,
            )
        ).all()
        assert len(linked) == count_rows(db_session, model, snapshot.id)
        assert all(record.source_key == raw.row_hash for record, raw in linked)
        assert all(raw.organization_id == golden_project.organization_id for _, raw in linked)
        assert all(raw.project_id == golden_project.id for _, raw in linked)

    progress = db_session.scalar(
        select(GovernedProgressRecord).where(
            GovernedProgressRecord.organization_id == golden_project.organization_id,
            GovernedProgressRecord.project_id == golden_project.id,
            GovernedProgressRecord.dataset_snapshot_id == snapshot.id,
            GovernedProgressRecord.progress_id == "PRG-50-4",
        )
    )
    contradictory = db_session.scalar(
        select(GovernedScheduleActivityRecord).where(
            GovernedScheduleActivityRecord.organization_id == golden_project.organization_id,
            GovernedScheduleActivityRecord.project_id == golden_project.id,
            GovernedScheduleActivityRecord.dataset_snapshot_id == snapshot.id,
            GovernedScheduleActivityRecord.activity_id == "A9990",
        )
    )
    source = db_session.scalar(
        select(SourceFileRecord).where(
            SourceFileRecord.organization_id == golden_project.organization_id,
            SourceFileRecord.project_id == golden_project.id,
            SourceFileRecord.id == snapshot.source_file_id,
        )
    )

    assert progress.actual_progress == Decimal("1.0800")
    assert contradictory.actual_progress == Decimal("1.1500")
    assert contradictory.actual_start == date(2026, 6, 20)
    assert contradictory.actual_finish == date(2026, 6, 10)
    assert storage.exists(source.storage_key)


def test_duplicate_business_ids_with_distinct_source_keys_remain_canonical(
    snapshot_service,
    golden_bytes,
    golden_project,
    db_session,
):
    def duplicate_budget_id(book) -> None:
        first_budget_id = book["Budget"].cell(row=4, column=1).value
        book["Budget"].cell(row=5, column=1, value=first_budget_id)

    duplicated = mutate_workbook(golden_bytes, duplicate_budget_id)
    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "duplicate-budget-id.xlsx",
        XLSX_MIME,
        duplicated,
    )
    records = db_session.scalars(
        select(GovernedBudgetRecord)
        .where(
            GovernedBudgetRecord.organization_id == golden_project.organization_id,
            GovernedBudgetRecord.project_id == golden_project.id,
            GovernedBudgetRecord.dataset_snapshot_id == snapshot.id,
            GovernedBudgetRecord.budget_id == "BUD-001",
        )
        .order_by(GovernedBudgetRecord.raw_row_id)
    ).all()
    raw_rows = db_session.scalars(
        select(GovernedRawRowRecord)
        .where(
            GovernedRawRowRecord.organization_id == golden_project.organization_id,
            GovernedRawRowRecord.project_id == golden_project.id,
            GovernedRawRowRecord.dataset_snapshot_id == snapshot.id,
            GovernedRawRowRecord.id.in_([record.raw_row_id for record in records]),
        )
        .order_by(GovernedRawRowRecord.source_row_number)
    ).all()

    assert snapshot.row_count_canonical == 149
    assert len(records) == 2
    assert len({record.source_key for record in records}) == 2
    assert len({record.raw_row_id for record in records}) == 2
    assert [raw.source_row_number for raw in raw_rows] == [4, 5]


def test_ingestion_persists_mapping_profile_definition_and_hash(
    snapshot_service,
    mapping_profile,
    golden_bytes,
    golden_project,
    db_session,
):
    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )
    persisted = db_session.scalar(
        select(GovernedMappingProfileVersionRecord).where(
            GovernedMappingProfileVersionRecord.id == snapshot.mapping_profile_version_id,
            GovernedMappingProfileVersionRecord.version == "0.1",
            GovernedMappingProfileVersionRecord.sha256
            == "1332b574985e8989c7b094a7ce99c11476defa9874d8aba4d0d874e46775497f",
        )
    )

    assert persisted is not None
    assert persisted.definition == mapping_profile.model_dump(mode="json")


def test_wbs_parent_and_fact_links_resolve_within_snapshot(
    snapshot_service,
    golden_bytes,
    golden_project,
    db_session,
):
    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )
    parent = db_session.scalar(
        select(GovernedWBSNodeRecord).where(
            GovernedWBSNodeRecord.organization_id == golden_project.organization_id,
            GovernedWBSNodeRecord.project_id == golden_project.id,
            GovernedWBSNodeRecord.dataset_snapshot_id == snapshot.id,
            GovernedWBSNodeRecord.wbs_code == "1.0",
        )
    )
    child = db_session.scalar(
        select(GovernedWBSNodeRecord).where(
            GovernedWBSNodeRecord.organization_id == golden_project.organization_id,
            GovernedWBSNodeRecord.project_id == golden_project.id,
            GovernedWBSNodeRecord.dataset_snapshot_id == snapshot.id,
            GovernedWBSNodeRecord.wbs_code == "1.1",
        )
    )
    budget = db_session.scalar(
        select(GovernedBudgetRecord).where(
            GovernedBudgetRecord.organization_id == golden_project.organization_id,
            GovernedBudgetRecord.project_id == golden_project.id,
            GovernedBudgetRecord.dataset_snapshot_id == snapshot.id,
            GovernedBudgetRecord.budget_id == "BUD-001",
        )
    )
    raw = db_session.scalar(
        select(GovernedRawRowRecord).where(
            GovernedRawRowRecord.organization_id == golden_project.organization_id,
            GovernedRawRowRecord.project_id == golden_project.id,
            GovernedRawRowRecord.dataset_snapshot_id == snapshot.id,
            GovernedRawRowRecord.id == budget.raw_row_id,
        )
    )

    assert child.parent_id == parent.id
    assert budget.wbs_node_id == child.id
    assert raw.source_sheet == "Budget"
    assert raw.source_row_number == 4


def test_database_failure_rolls_back_snapshot_and_deletes_source(
    snapshot_service,
    golden_bytes,
    golden_project,
    db_session,
    storage,
    monkeypatch,
):
    def fail_canonical_write(*args, **kwargs):
        raise RuntimeError("injected canonical database failure")

    monkeypatch.setattr(SnapshotRepository, "persist_canonical_rows", fail_canonical_write)

    with pytest.raises(RuntimeError, match="injected canonical database failure"):
        snapshot_service.ingest(
            golden_project.organization_id,
            golden_project.id,
            "golden.xlsx",
            XLSX_MIME,
            golden_bytes,
        )

    assert stored_files(storage) == []
    assert count_scoped(
        db_session, GovernedDatasetSnapshotRecord, golden_project.organization_id, golden_project.id
    ) == 0
    assert count_scoped(
        db_session, SourceFileRecord, golden_project.organization_id, golden_project.id
    ) == 0


def test_storage_failure_creates_no_snapshot(
    session_factory,
    mapping_profile,
    golden_bytes,
    golden_project,
    db_session,
    tmp_path,
):
    class FailingPutStorage(LocalFileStorage):
        def put(self, organization_id, project_id, filename, data):
            raise OSError("injected storage failure")

    service = SnapshotIngestionService(
        session_factory,
        FailingPutStorage(tmp_path),
        mapping_profile,
    )

    with pytest.raises(OSError, match="injected storage failure"):
        service.ingest(
            golden_project.organization_id,
            golden_project.id,
            "golden.xlsx",
            XLSX_MIME,
            golden_bytes,
        )

    assert count_scoped(
        db_session, GovernedDatasetSnapshotRecord, golden_project.organization_id, golden_project.id
    ) == 0
    assert count_scoped(
        db_session, SourceFileRecord, golden_project.organization_id, golden_project.id
    ) == 0


def test_missing_source_before_completion_rolls_back_snapshot(
    session_factory,
    mapping_profile,
    golden_bytes,
    golden_project,
    db_session,
    tmp_path,
):
    class VanishingStorage(LocalFileStorage):
        def put(self, organization_id, project_id, filename, data):
            stored = super().put(organization_id, project_id, filename, data)
            self.delete(stored.key)
            return stored

    storage = VanishingStorage(tmp_path)
    service = SnapshotIngestionService(session_factory, storage, mapping_profile)

    with pytest.raises(ControlCheckApplicationError) as exc_info:
        service.ingest(
            golden_project.organization_id,
            golden_project.id,
            "golden.xlsx",
            XLSX_MIME,
            golden_bytes,
        )

    assert exc_info.value.code == "source_object_missing"
    assert stored_files(storage) == []
    assert count_scoped(
        db_session, GovernedDatasetSnapshotRecord, golden_project.organization_id, golden_project.id
    ) == 0
    assert count_scoped(
        db_session, SourceFileRecord, golden_project.organization_id, golden_project.id
    ) == 0


def test_post_commit_detach_failure_keeps_committed_snapshot_source(
    snapshot_service,
    golden_bytes,
    golden_project,
    db_session,
    storage,
    monkeypatch,
):
    def fail_after_commit(session, snapshot):
        raise RuntimeError("injected post-commit detach failure")

    monkeypatch.setattr(
        SnapshotIngestionService,
        "_detached",
        staticmethod(fail_after_commit),
    )

    with pytest.raises(RuntimeError, match="injected post-commit detach failure"):
        snapshot_service.ingest(
            golden_project.organization_id,
            golden_project.id,
            "golden.xlsx",
            XLSX_MIME,
            golden_bytes,
        )

    snapshot = db_session.scalar(
        select(GovernedDatasetSnapshotRecord).where(
            GovernedDatasetSnapshotRecord.organization_id == golden_project.organization_id,
            GovernedDatasetSnapshotRecord.project_id == golden_project.id,
        )
    )
    source = db_session.scalar(
        select(SourceFileRecord).where(
            SourceFileRecord.organization_id == golden_project.organization_id,
            SourceFileRecord.project_id == golden_project.id,
            SourceFileRecord.id == snapshot.source_file_id,
        )
    )
    assert snapshot.status == "validated"
    assert storage.exists(source.storage_key)


def test_post_server_commit_exception_reconciles_snapshot_and_preserves_source(
    snapshot_service,
    session_factory,
    golden_bytes,
    golden_project,
    storage,
    monkeypatch,
):
    session_class = session_factory.class_
    original_commit = session_class.commit
    injected = False

    def commit_then_raise(session):
        nonlocal injected
        completed_snapshot = any(
            isinstance(item, GovernedDatasetSnapshotRecord)
            and item.status in {"validated", "validated_with_errors"}
            for item in session.identity_map.values()
        )
        original_commit(session)
        if completed_snapshot and not injected:
            injected = True
            raise RuntimeError("injected post-server-commit transport failure")

    monkeypatch.setattr(session_class, "commit", commit_then_raise)

    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )

    assert injected
    assert snapshot.status == "validated"
    with session_factory() as session:
        persisted = session.get(GovernedDatasetSnapshotRecord, snapshot.id)
        source = session.get(SourceFileRecord, persisted.source_file_id)
    assert persisted.status == "validated"
    assert storage.exists(source.storage_key)


def test_project_code_mismatch_is_stored_as_source_identity(
    snapshot_service,
    golden_bytes,
    golden_project,
    db_session,
    storage,
):
    mismatched = mutate_workbook(
        golden_bytes,
        lambda book: replace_project_id(book, "SOME-OTHER-PROJECT"),
    )

    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "mismatched.xlsx",
        XLSX_MIME,
        mismatched,
    )

    assert snapshot.status in {"validated", "validated_with_errors"}
    assert snapshot.source_project_id == "SOME-OTHER-PROJECT"
    assert stored_files(storage)
    assert count_scoped(
        db_session, GovernedDatasetSnapshotRecord, golden_project.organization_id, golden_project.id
    ) == 1
    assert count_scoped(
        db_session, SourceFileRecord, golden_project.organization_id, golden_project.id
    ) == 1


def test_missing_source_project_metadata_uses_selected_project_context(
    snapshot_service,
    golden_bytes,
    golden_project,
):
    def remove_project_metadata(book):
        sheet = book["Project_Info"]
        for row in range(sheet.max_row, 1, -1):
            if sheet.cell(row=row, column=1).value in {"project_id", "project_name"}:
                sheet.delete_rows(row)

    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "missing-project-metadata.xlsx",
        XLSX_MIME,
        mutate_workbook(golden_bytes, remove_project_metadata),
    )

    assert snapshot.source_project_id == golden_project.code
    assert snapshot.source_project_name == golden_project.name


def test_missing_sheet_marks_only_that_domain_blocked_and_persists_healthy_rows(
    snapshot_service,
    golden_bytes,
    golden_project,
    db_session,
):
    missing_budget = mutate_workbook(golden_bytes, lambda book: book.remove(book["Budget"]))

    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "missing-budget.xlsx",
        XLSX_MIME,
        missing_budget,
    )
    statuses = {
        row.domain: row
        for row in db_session.scalars(
            select(GovernedDatasetDomainStatusRecord).where(
                GovernedDatasetDomainStatusRecord.organization_id == golden_project.organization_id,
                GovernedDatasetDomainStatusRecord.project_id == golden_project.id,
                GovernedDatasetDomainStatusRecord.dataset_snapshot_id == snapshot.id,
            )
        )
    }

    assert snapshot.status == "validated_with_errors"
    assert snapshot.row_count_raw == 140
    assert snapshot.row_count_canonical == 140
    assert statuses["budget"].status == "blocked"
    assert statuses["budget"].row_count_raw == 0
    assert statuses["budget"].error_count == 1
    assert statuses["actual_cost"].status == "valid"
    assert count_rows(db_session, GovernedActualCostRecord, snapshot.id) == 73
    assert count_rows(db_session, GovernedScheduleActivityRecord, snapshot.id) == 13


def test_wbs_blocking_persists_dependency_reason_for_dependent_domains(
    snapshot_service,
    golden_bytes,
    golden_project,
    db_session,
):
    missing_wbs = mutate_workbook(golden_bytes, lambda book: book.remove(book["WBS"]))

    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "missing-wbs.xlsx",
        XLSX_MIME,
        missing_wbs,
    )
    statuses = {
        row.domain: row
        for row in db_session.scalars(
            select(GovernedDatasetDomainStatusRecord).where(
                GovernedDatasetDomainStatusRecord.organization_id == golden_project.organization_id,
                GovernedDatasetDomainStatusRecord.project_id == golden_project.id,
                GovernedDatasetDomainStatusRecord.dataset_snapshot_id == snapshot.id,
            )
        )
    }
    expected_issue = {
        "code": "blocked_by_wbs",
        "message": "Domain is blocked because the WBS domain is blocked",
        "dependency_domain": "wbs",
        "severity": "error",
    }

    assert statuses["wbs"].status == "blocked"
    for domain in ("budget", "actual_cost", "commitments", "schedule", "progress"):
        assert statuses[domain].status == "blocked"
        assert statuses[domain].validation_summary["dependency_issues"] == [expected_issue]
        assert statuses[domain].error_count >= 1


def test_hard_invalid_row_is_raw_only_while_warning_row_is_canonical(
    snapshot_service,
    golden_bytes,
    golden_project,
    db_session,
):
    def add_invalid_and_warning(book) -> None:
        wbs_name = book["WBS"].cell(row=4, column=2).value
        book["WBS"].cell(row=4, column=2, value=f"  {wbs_name}  ")
        book["Progress"].cell(row=4, column=5, value="not-a-number")

    changed = mutate_workbook(golden_bytes, add_invalid_and_warning)
    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "row-validation.xlsx",
        XLSX_MIME,
        changed,
    )
    invalid = db_session.scalar(
        select(GovernedRawRowRecord).where(
            GovernedRawRowRecord.organization_id == golden_project.organization_id,
            GovernedRawRowRecord.project_id == golden_project.id,
            GovernedRawRowRecord.dataset_snapshot_id == snapshot.id,
            GovernedRawRowRecord.domain == "progress",
            GovernedRawRowRecord.source_row_number == 4,
        )
    )
    warning = db_session.scalar(
        select(GovernedRawRowRecord).where(
            GovernedRawRowRecord.organization_id == golden_project.organization_id,
            GovernedRawRowRecord.project_id == golden_project.id,
            GovernedRawRowRecord.dataset_snapshot_id == snapshot.id,
            GovernedRawRowRecord.domain == "wbs",
            GovernedRawRowRecord.source_row_number == 4,
        )
    )

    assert snapshot.status == "validated_with_errors"
    assert snapshot.row_count_raw == 149
    assert snapshot.row_count_canonical == 148
    assert invalid.validation_status == "invalid"
    assert invalid.validation_errors[0]["code"] == "invalid_decimal"
    assert db_session.scalar(
        select(GovernedProgressRecord.id).where(
            GovernedProgressRecord.dataset_snapshot_id == snapshot.id,
            GovernedProgressRecord.raw_row_id == invalid.id,
        )
    ) is None
    assert warning.validation_status == "warning"
    assert db_session.scalar(
        select(GovernedWBSNodeRecord.id).where(
            GovernedWBSNodeRecord.dataset_snapshot_id == snapshot.id,
            GovernedWBSNodeRecord.raw_row_id == warning.id,
        )
    ) is not None


def test_snapshot_queries_are_tenant_and_project_scoped(
    snapshot_service,
    session_factory,
    golden_bytes,
    golden_project,
):
    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )
    with session_factory() as session:
        other_organization = OrganizationRecord(
            name="Other organization",
            slug=f"other-org-{uuid4().hex}",
        )
        session.add(other_organization)
        session.flush()
        other_project = ProjectRecord(
            organization_id=golden_project.organization_id,
            code=f"OTHER-{uuid4().hex[:8]}",
            name="Other project",
            currency="IDR",
        )
        session.add(other_project)
        session.flush()
        repository = SnapshotRepository(session)

        assert repository.get_scoped(
            golden_project.organization_id, golden_project.id, snapshot.id
        ).id == snapshot.id
        assert repository.list_scoped(
            golden_project.organization_id, golden_project.id
        )[0].id == snapshot.id
        assert repository.get_scoped(
            other_organization.id, golden_project.id, snapshot.id
        ) is None
        assert repository.get_scoped(
            golden_project.organization_id, other_project.id, snapshot.id
        ) is None
        assert repository.list_scoped(other_organization.id, golden_project.id) == []
        assert repository.list_scoped(
            golden_project.organization_id, other_project.id
        ) == []
        assert repository.find_duplicate(
            other_organization.id, golden_project.id, snapshot.dedupe_key
        ) is None


def test_completed_snapshot_cannot_accept_more_canonical_facts(
    snapshot_service,
    session_factory,
    mapping_profile,
    golden_bytes,
    golden_project,
):
    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )
    extracted = extract_workbook(golden_bytes, mapping_profile)
    mapped = map_extracted_workbook(extracted, mapping_profile)

    with session_factory() as session:
        repository = SnapshotRepository(session)
        with pytest.raises(SnapshotImmutableError):
            repository.persist_canonical_rows(
                golden_project.organization_id,
                golden_project.id,
                snapshot.id,
                mapped,
                {},
            )
        with pytest.raises(SnapshotImmutableError):
            repository.fail(
                golden_project.organization_id,
                golden_project.id,
                snapshot.id,
                "late_failure",
                "Completed snapshots cannot fail",
            )


def test_failed_snapshot_releases_dedupe_key_and_normal_retry_succeeds(
    session_factory,
    storage,
    mapping_profile,
    golden_bytes,
    golden_project,
    monkeypatch,
):
    dedupe_key = golden_project.organization_id.hex + golden_project.id.hex
    monkeypatch.setattr(
        ingestion_service_module,
        "_dedupe_key",
        lambda *args: dedupe_key,
    )
    with session_factory() as session:
        repository, snapshot, stored, _, _ = begin_ingesting_snapshot(
            session,
            storage,
            mapping_profile,
            golden_bytes,
            golden_project,
            dedupe_key,
        )
        failed = repository.fail(
            golden_project.organization_id,
            golden_project.id,
            snapshot.id,
            "injected_failure",
            "The first import failed safely",
        )
        session.commit()

    assert failed.status == "failed"
    assert failed.dedupe_key is None
    assert storage.exists(stored.key)

    retried = SnapshotIngestionService(
        session_factory,
        storage,
        mapping_profile,
    ).ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden-retry.xlsx",
        XLSX_MIME,
        golden_bytes,
    )

    assert retried.status == "validated"
    assert retried.id != failed.id
    assert retried.dedupe_key == dedupe_key


def test_fail_atomically_removes_partial_rows_facts_and_domain_statuses(
    session_factory,
    storage,
    mapping_profile,
    golden_bytes,
    golden_project,
):
    with session_factory() as session:
        repository, snapshot, stored, extracted, mapped = begin_ingesting_snapshot(
            session,
            storage,
            mapping_profile,
            golden_bytes,
            golden_project,
            golden_project.id.hex + golden_project.organization_id.hex,
        )
        raw_rows = repository.persist_raw_rows(
            golden_project.organization_id,
            golden_project.id,
            snapshot.id,
            extracted,
            mapped,
        )
        repository.persist_canonical_rows(
            golden_project.organization_id,
            golden_project.id,
            snapshot.id,
            mapped,
            raw_rows,
        )
        session.add(
            GovernedDatasetDomainStatusRecord(
                organization_id=golden_project.organization_id,
                project_id=golden_project.id,
                dataset_snapshot_id=snapshot.id,
                domain="wbs",
                status="valid",
                row_count_raw=12,
                row_count_canonical=12,
                error_count=0,
                warning_count=0,
                validation_summary={},
            )
        )
        batch = session.scalar(
            select(GovernedImportBatchRecord).where(
                GovernedImportBatchRecord.organization_id == golden_project.organization_id,
                GovernedImportBatchRecord.project_id == golden_project.id,
                GovernedImportBatchRecord.dataset_snapshot_id == snapshot.id,
                GovernedImportBatchRecord.id == snapshot.import_batch_id,
            )
        )
        batch.rows_read = 149
        batch.rows_valid = 149
        snapshot.row_count_raw = 149
        snapshot.row_count_canonical = 149
        session.flush()

        failed = repository.fail(
            golden_project.organization_id,
            golden_project.id,
            snapshot.id,
            "canonical_write_failed",
            "Canonical persistence failed safely",
            {"stage": "canonical"},
        )
        failed_id = failed.id
        batch_id = batch.id
        session.commit()

    with session_factory() as session:
        failed = SnapshotRepository(session).get_scoped(
            golden_project.organization_id,
            golden_project.id,
            failed_id,
        )
        batch = session.scalar(
            select(GovernedImportBatchRecord).where(
                GovernedImportBatchRecord.organization_id == golden_project.organization_id,
                GovernedImportBatchRecord.project_id == golden_project.id,
                GovernedImportBatchRecord.dataset_snapshot_id == failed_id,
                GovernedImportBatchRecord.id == batch_id,
            )
        )
        domain_count = session.scalar(
            select(func.count()).select_from(GovernedDatasetDomainStatusRecord).where(
                GovernedDatasetDomainStatusRecord.organization_id == golden_project.organization_id,
                GovernedDatasetDomainStatusRecord.project_id == golden_project.id,
                GovernedDatasetDomainStatusRecord.dataset_snapshot_id == failed_id,
            )
        )
        raw_count = count_rows(session, GovernedRawRowRecord, failed_id)
        canonical_count = sum(count_rows(session, model, failed_id) for model in CANONICAL_MODELS)

    assert failed.status == "failed"
    assert failed.dedupe_key is None
    assert failed.row_count_raw == 0
    assert failed.row_count_canonical == 0
    assert batch.status == "failed"
    assert batch.rows_read == 0
    assert batch.rows_valid == 0
    assert batch.rows_warning == 0
    assert batch.rows_rejected == 0
    assert batch.safe_error_code == "canonical_write_failed"
    assert batch.safe_error_message == "Canonical persistence failed safely"
    assert batch.error_summary == {"stage": "canonical"}
    assert domain_count == 0
    assert raw_count == 0
    assert canonical_count == 0
    assert storage.exists(stored.key)


def test_repository_refuses_completion_without_all_six_domain_states(
    session_factory,
    storage,
    mapping_profile,
    golden_bytes,
    golden_project,
):
    with session_factory() as session:
        repository, snapshot, _, _, _ = begin_ingesting_snapshot(
            session,
            storage,
            mapping_profile,
            golden_bytes,
            golden_project,
            golden_project.organization_id.hex + golden_project.id.hex,
        )

        with pytest.raises(ValueError, match="all six governed domain states"):
            repository.complete(
                golden_project.organization_id,
                golden_project.id,
                snapshot.id,
                status="validated",
                row_count_raw=0,
                row_count_canonical=0,
            )


def test_duplicate_race_returns_winner_and_deletes_losing_file(
    snapshot_service,
    golden_bytes,
    golden_project,
    storage,
    monkeypatch,
):
    winner = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )
    original_find_duplicate = SnapshotRepository.find_duplicate
    calls = 0

    def hide_winner_once(repository, organization_id, project_id, dedupe_key):
        nonlocal calls
        calls += 1
        if calls == 1:
            return None
        return original_find_duplicate(repository, organization_id, project_id, dedupe_key)

    monkeypatch.setattr(SnapshotRepository, "find_duplicate", hide_winner_once)
    before = stored_files(storage)

    raced = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "renamed.xlsx",
        XLSX_MIME,
        golden_bytes,
    )

    assert winner.outcome == "created"
    assert raced.outcome == "deduplicated"
    assert raced.snapshot.id == winner.snapshot.id
    assert stored_files(storage) == before
    assert calls == 2
