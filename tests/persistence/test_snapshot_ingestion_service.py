from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from uuid import uuid4

import openpyxl
import pytest
from alembic import command
from sqlalchemy import and_, func, select

from controlcheck.errors import ControlCheckApplicationError
from controlcheck.ingestion.extractor import extract_workbook
from controlcheck.ingestion.mapper import map_extracted_workbook
from controlcheck.ingestion.profile import load_mapping_profile
from controlcheck.ingestion.service import SnapshotIngestionService
from controlcheck.persistence.database import create_session_factory
from controlcheck.persistence.ingestion_repositories import (
    SnapshotImmutableError,
    SnapshotRepository,
)
from controlcheck.persistence.models import (
    CanonicalActualCostRecord,
    CanonicalBudgetRecord,
    CanonicalCommitmentRecord,
    DatasetDomainStatusRecord,
    DatasetSnapshotRecord,
    OrganizationRecord,
    ProgressRecordRecord,
    ProjectRecord,
    RawRowRecord,
    ScheduleActivityRecord,
    SourceFileRecord,
    WBSNodeRecord,
)
from controlcheck.storage import LocalFileStorage


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CANONICAL_MODELS = (
    WBSNodeRecord,
    CanonicalBudgetRecord,
    CanonicalActualCostRecord,
    CanonicalCommitmentRecord,
    ScheduleActivityRecord,
    ProgressRecordRecord,
)


@pytest.fixture()
def session_factory(alembic_config, postgres_url):
    command.upgrade(alembic_config, "head")
    return create_session_factory(postgres_url)


@pytest.fixture()
def db_session(session_factory):
    with session_factory() as session:
        yield session


@pytest.fixture()
def golden_project(session_factory) -> ProjectRecord:
    with session_factory() as session:
        organization = OrganizationRecord(
            name="Snapshot ingestion organization",
            slug=f"snapshot-ingestion-{uuid4().hex}",
        )
        session.add(organization)
        session.flush()
        project = ProjectRecord(
            organization_id=organization.id,
            code="PRJ-CCAI-001",
            name="Golden project",
            currency="IDR",
        )
        session.add(project)
        session.commit()
        return project


@pytest.fixture(scope="session")
def golden_bytes(project_root: Path) -> bytes:
    return (project_root / "data" / "ControlCheck_AI_Golden_Positive_Dataset_v0.2.xlsx").read_bytes()


@pytest.fixture()
def mapping_profile(project_root: Path):
    return load_mapping_profile(project_root / "data" / "controlcheck_mapping_profile_v0.1.json")


@pytest.fixture()
def storage(tmp_path: Path):
    return LocalFileStorage(tmp_path)


@pytest.fixture()
def snapshot_service(session_factory, storage, mapping_profile):
    return SnapshotIngestionService(
        session_factory=session_factory,
        storage=storage,
        mapping_profile=mapping_profile,
    )


def count_rows(db_session, model, snapshot_id) -> int:
    return db_session.scalar(
        select(func.count()).select_from(model).where(model.dataset_snapshot_id == snapshot_id)
    )


def count_scoped(db_session, model, organization_id, project_id) -> int:
    return db_session.scalar(
        select(func.count()).select_from(model).where(
            model.organization_id == organization_id,
            model.project_id == project_id,
        )
    )


def stored_files(storage: LocalFileStorage) -> list[Path]:
    return sorted(path for path in storage.root.rglob("*") if path.is_file())


def mutate_workbook(data: bytes, mutation) -> bytes:
    book = openpyxl.load_workbook(BytesIO(data))
    try:
        mutation(book)
        output = BytesIO()
        book.save(output)
        return output.getvalue()
    finally:
        book.close()


def replace_project_id(book, value: str) -> None:
    sheet = book["Project_Info"]
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == "project_id":
            sheet.cell(row=row, column=2, value=value)
            return
    raise AssertionError("Golden workbook has no project_id metadata row")


def test_golden_ingestion_persists_raw_and_canonical_rows(
    snapshot_service, golden_bytes, golden_project, db_session
):
    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )

    assert snapshot.status == "validated"
    assert snapshot.row_count_raw == 149
    assert snapshot.row_count_canonical == 149
    assert count_rows(db_session, RawRowRecord, snapshot.id) == 149
    assert sum(count_rows(db_session, model, snapshot.id) for model in CANONICAL_MODELS) == 149


def test_duplicate_upload_returns_existing_snapshot(
    snapshot_service, golden_bytes, golden_project, storage
):
    first = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )
    second = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "renamed.xlsx",
        XLSX_MIME,
        golden_bytes,
    )

    assert second.id == first.id
    assert len(stored_files(storage)) == 1


def test_force_new_creates_distinct_snapshot(snapshot_service, golden_bytes, golden_project):
    first = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )
    forced = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
        force_new=True,
    )

    assert forced.id != first.id
    assert forced.dedupe_key is None


def test_golden_rows_keep_exact_lineage_and_rule_detectable_anomalies(
    snapshot_service, golden_bytes, golden_project, db_session, storage
):
    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )

    for model in CANONICAL_MODELS:
        linked = db_session.execute(
            select(model, RawRowRecord)
            .join(
                RawRowRecord,
                and_(
                    RawRowRecord.dataset_snapshot_id == model.dataset_snapshot_id,
                    RawRowRecord.id == model.raw_row_id,
                ),
            )
            .where(
                model.organization_id == golden_project.organization_id,
                model.project_id == golden_project.id,
                model.dataset_snapshot_id == snapshot.id,
            )
        ).all()
        assert len(linked) == count_rows(db_session, model, snapshot.id)
        assert all(record.source_key == raw.row_hash for record, raw in linked)
        assert all(raw.organization_id == golden_project.organization_id for _, raw in linked)
        assert all(raw.project_id == golden_project.id for _, raw in linked)

    progress = db_session.scalar(
        select(ProgressRecordRecord).where(
            ProgressRecordRecord.organization_id == golden_project.organization_id,
            ProgressRecordRecord.project_id == golden_project.id,
            ProgressRecordRecord.dataset_snapshot_id == snapshot.id,
            ProgressRecordRecord.progress_id == "PRG-50-4",
        )
    )
    contradictory = db_session.scalar(
        select(ScheduleActivityRecord).where(
            ScheduleActivityRecord.organization_id == golden_project.organization_id,
            ScheduleActivityRecord.project_id == golden_project.id,
            ScheduleActivityRecord.dataset_snapshot_id == snapshot.id,
            ScheduleActivityRecord.activity_id == "A9990",
        )
    )
    source = db_session.scalar(
        select(SourceFileRecord).where(
            SourceFileRecord.organization_id == golden_project.organization_id,
            SourceFileRecord.project_id == golden_project.id,
            SourceFileRecord.id == snapshot.source_file_id,
        )
    )

    assert progress.actual_progress == Decimal("1.0800")
    assert contradictory.actual_progress == Decimal("1.1500")
    assert contradictory.actual_start == date(2026, 6, 20)
    assert contradictory.actual_finish == date(2026, 6, 10)
    assert storage.exists(source.storage_key)


def test_database_failure_rolls_back_snapshot_and_deletes_source(
    snapshot_service,
    golden_bytes,
    golden_project,
    db_session,
    storage,
    monkeypatch,
):
    def fail_canonical_write(*args, **kwargs):
        raise RuntimeError("injected canonical database failure")

    monkeypatch.setattr(SnapshotRepository, "persist_canonical_rows", fail_canonical_write)

    with pytest.raises(RuntimeError, match="injected canonical database failure"):
        snapshot_service.ingest(
            golden_project.organization_id,
            golden_project.id,
            "golden.xlsx",
            XLSX_MIME,
            golden_bytes,
        )

    assert stored_files(storage) == []
    assert count_scoped(
        db_session, DatasetSnapshotRecord, golden_project.organization_id, golden_project.id
    ) == 0
    assert count_scoped(
        db_session, SourceFileRecord, golden_project.organization_id, golden_project.id
    ) == 0


def test_storage_failure_creates_no_snapshot(
    session_factory,
    mapping_profile,
    golden_bytes,
    golden_project,
    db_session,
    tmp_path,
):
    class FailingPutStorage(LocalFileStorage):
        def put(self, organization_id, project_id, filename, data):
            raise OSError("injected storage failure")

    service = SnapshotIngestionService(
        session_factory,
        FailingPutStorage(tmp_path),
        mapping_profile,
    )

    with pytest.raises(OSError, match="injected storage failure"):
        service.ingest(
            golden_project.organization_id,
            golden_project.id,
            "golden.xlsx",
            XLSX_MIME,
            golden_bytes,
        )

    assert count_scoped(
        db_session, DatasetSnapshotRecord, golden_project.organization_id, golden_project.id
    ) == 0
    assert count_scoped(
        db_session, SourceFileRecord, golden_project.organization_id, golden_project.id
    ) == 0


def test_missing_source_before_completion_rolls_back_snapshot(
    session_factory,
    mapping_profile,
    golden_bytes,
    golden_project,
    db_session,
    tmp_path,
):
    class VanishingStorage(LocalFileStorage):
        def put(self, organization_id, project_id, filename, data):
            stored = super().put(organization_id, project_id, filename, data)
            self.delete(stored.key)
            return stored

    storage = VanishingStorage(tmp_path)
    service = SnapshotIngestionService(session_factory, storage, mapping_profile)

    with pytest.raises(ControlCheckApplicationError) as exc_info:
        service.ingest(
            golden_project.organization_id,
            golden_project.id,
            "golden.xlsx",
            XLSX_MIME,
            golden_bytes,
        )

    assert exc_info.value.code == "source_object_missing"
    assert stored_files(storage) == []
    assert count_scoped(
        db_session, DatasetSnapshotRecord, golden_project.organization_id, golden_project.id
    ) == 0
    assert count_scoped(
        db_session, SourceFileRecord, golden_project.organization_id, golden_project.id
    ) == 0


def test_project_code_mismatch_stores_nothing(
    snapshot_service,
    golden_bytes,
    golden_project,
    db_session,
    storage,
):
    mismatched = mutate_workbook(
        golden_bytes,
        lambda book: replace_project_id(book, "SOME-OTHER-PROJECT"),
    )

    with pytest.raises(ControlCheckApplicationError) as exc_info:
        snapshot_service.ingest(
            golden_project.organization_id,
            golden_project.id,
            "mismatched.xlsx",
            XLSX_MIME,
            mismatched,
        )

    assert exc_info.value.code == "workbook_project_mismatch"
    assert stored_files(storage) == []
    assert count_scoped(
        db_session, DatasetSnapshotRecord, golden_project.organization_id, golden_project.id
    ) == 0
    assert count_scoped(
        db_session, SourceFileRecord, golden_project.organization_id, golden_project.id
    ) == 0


def test_missing_sheet_marks_only_that_domain_blocked_and_persists_healthy_rows(
    snapshot_service,
    golden_bytes,
    golden_project,
    db_session,
):
    missing_budget = mutate_workbook(golden_bytes, lambda book: book.remove(book["Budget"]))

    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "missing-budget.xlsx",
        XLSX_MIME,
        missing_budget,
    )
    statuses = {
        row.domain: row
        for row in db_session.scalars(
            select(DatasetDomainStatusRecord).where(
                DatasetDomainStatusRecord.organization_id == golden_project.organization_id,
                DatasetDomainStatusRecord.project_id == golden_project.id,
                DatasetDomainStatusRecord.dataset_snapshot_id == snapshot.id,
            )
        )
    }

    assert snapshot.status == "validated_with_errors"
    assert snapshot.row_count_raw == 140
    assert snapshot.row_count_canonical == 140
    assert statuses["budget"].status == "blocked"
    assert statuses["budget"].row_count_raw == 0
    assert statuses["budget"].error_count == 1
    assert statuses["actual_cost"].status == "valid"
    assert count_rows(db_session, CanonicalActualCostRecord, snapshot.id) == 73
    assert count_rows(db_session, ScheduleActivityRecord, snapshot.id) == 13


def test_hard_invalid_row_is_raw_only_while_warning_row_is_canonical(
    snapshot_service,
    golden_bytes,
    golden_project,
    db_session,
):
    def add_invalid_and_warning(book) -> None:
        wbs_name = book["WBS"].cell(row=4, column=2).value
        book["WBS"].cell(row=4, column=2, value=f"  {wbs_name}  ")
        book["Progress"].cell(row=4, column=5, value="not-a-number")

    changed = mutate_workbook(golden_bytes, add_invalid_and_warning)
    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "row-validation.xlsx",
        XLSX_MIME,
        changed,
    )
    invalid = db_session.scalar(
        select(RawRowRecord).where(
            RawRowRecord.organization_id == golden_project.organization_id,
            RawRowRecord.project_id == golden_project.id,
            RawRowRecord.dataset_snapshot_id == snapshot.id,
            RawRowRecord.domain == "progress",
            RawRowRecord.source_row_number == 4,
        )
    )
    warning = db_session.scalar(
        select(RawRowRecord).where(
            RawRowRecord.organization_id == golden_project.organization_id,
            RawRowRecord.project_id == golden_project.id,
            RawRowRecord.dataset_snapshot_id == snapshot.id,
            RawRowRecord.domain == "wbs",
            RawRowRecord.source_row_number == 4,
        )
    )

    assert snapshot.status == "validated_with_errors"
    assert snapshot.row_count_raw == 149
    assert snapshot.row_count_canonical == 148
    assert invalid.validation_status == "invalid"
    assert invalid.validation_errors[0]["code"] == "invalid_decimal"
    assert db_session.scalar(
        select(ProgressRecordRecord.id).where(
            ProgressRecordRecord.dataset_snapshot_id == snapshot.id,
            ProgressRecordRecord.raw_row_id == invalid.id,
        )
    ) is None
    assert warning.validation_status == "warning"
    assert db_session.scalar(
        select(WBSNodeRecord.id).where(
            WBSNodeRecord.dataset_snapshot_id == snapshot.id,
            WBSNodeRecord.raw_row_id == warning.id,
        )
    ) is not None


def test_snapshot_queries_are_tenant_and_project_scoped(
    snapshot_service,
    session_factory,
    golden_bytes,
    golden_project,
):
    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )
    with session_factory() as session:
        other_organization = OrganizationRecord(
            name="Other organization",
            slug=f"other-org-{uuid4().hex}",
        )
        session.add(other_organization)
        session.flush()
        other_project = ProjectRecord(
            organization_id=golden_project.organization_id,
            code=f"OTHER-{uuid4().hex[:8]}",
            name="Other project",
            currency="IDR",
        )
        session.add(other_project)
        session.flush()
        repository = SnapshotRepository(session)

        assert repository.get_scoped(
            golden_project.organization_id, golden_project.id, snapshot.id
        ).id == snapshot.id
        assert repository.list_scoped(
            golden_project.organization_id, golden_project.id
        )[0].id == snapshot.id
        assert repository.get_scoped(
            other_organization.id, golden_project.id, snapshot.id
        ) is None
        assert repository.get_scoped(
            golden_project.organization_id, other_project.id, snapshot.id
        ) is None
        assert repository.list_scoped(other_organization.id, golden_project.id) == []
        assert repository.list_scoped(
            golden_project.organization_id, other_project.id
        ) == []
        assert repository.find_duplicate(
            other_organization.id, golden_project.id, snapshot.dedupe_key
        ) is None


def test_completed_snapshot_cannot_accept_more_canonical_facts(
    snapshot_service,
    session_factory,
    mapping_profile,
    golden_bytes,
    golden_project,
):
    snapshot = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )
    extracted = extract_workbook(golden_bytes, mapping_profile)
    mapped = map_extracted_workbook(extracted, mapping_profile)

    with session_factory() as session:
        repository = SnapshotRepository(session)
        with pytest.raises(SnapshotImmutableError):
            repository.persist_canonical_rows(
                golden_project.organization_id,
                golden_project.id,
                snapshot.id,
                mapped,
                {},
            )
        with pytest.raises(SnapshotImmutableError):
            repository.fail(
                golden_project.organization_id,
                golden_project.id,
                snapshot.id,
                "late_failure",
                "Completed snapshots cannot fail",
            )


def test_duplicate_race_returns_winner_and_deletes_losing_file(
    snapshot_service,
    golden_bytes,
    golden_project,
    storage,
    monkeypatch,
):
    winner = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "golden.xlsx",
        XLSX_MIME,
        golden_bytes,
    )
    original_find_duplicate = SnapshotRepository.find_duplicate
    calls = 0

    def hide_winner_once(repository, organization_id, project_id, dedupe_key):
        nonlocal calls
        calls += 1
        if calls == 1:
            return None
        return original_find_duplicate(repository, organization_id, project_id, dedupe_key)

    monkeypatch.setattr(SnapshotRepository, "find_duplicate", hide_winner_once)
    before = stored_files(storage)

    raced = snapshot_service.ingest(
        golden_project.organization_id,
        golden_project.id,
        "renamed.xlsx",
        XLSX_MIME,
        golden_bytes,
    )

    assert raced.id == winner.id
    assert stored_files(storage) == before
    assert calls == 2
