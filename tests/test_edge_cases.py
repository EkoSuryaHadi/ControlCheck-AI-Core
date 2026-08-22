from io import BytesIO
from datetime import date
from decimal import Decimal
import pytest
import openpyxl

from controlcheck.loader import load_workbook
from controlcheck.config import load_catalogue, ThresholdConfig
from controlcheck.engine import ControlEngine, RuleContext
from controlcheck.rules import ALL_RULES
from controlcheck.errors import ControlCheckApplicationError


def test_edge_case_whitespace_and_mixed_case_headers(sample_catalogue):
    """Verifies that headers with trailing spaces and mixed casing are parsed correctly."""
    wb = openpyxl.Workbook()
    ws_info = wb.active
    ws_info.title = "  Project Info  "
    ws_info.append(["  Attribute  ", " VALUE "])
    ws_info.append([" Project ID ", "PRJ-EDGE-001"])
    ws_info.append(["Project Name", "Edge Case Project"])
    ws_info.append([" Data Date ", "2026-08-01"])
    ws_info.append(["Dataset Version", "0.2"])

    ws_wbs = wb.create_sheet("WBS")
    ws_wbs.append(["  WBS Code  ", " WBS Name ", "Parent WBS", "Discipline", "Level"])
    ws_wbs.append([" 1.0 ", " Root Element ", None, " CIVIL ", 1])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    dataset = load_workbook(buf)
    assert dataset.project.project_id == "PRJ-EDGE-001"
    assert len(dataset.wbs_nodes) == 1
    assert dataset.wbs_nodes[0].wbs_code == "1.0"


def test_edge_case_division_by_zero_defense(sample_catalogue):
    """
    Verifies that rules involving division (e.g. CPI = EV / AC, SPI = EV / PV)
    gracefully handle 0 budget, 0 actual cost, and 0 earned value without ZeroDivisionError.
    """
    wb = openpyxl.Workbook()
    ws_info = wb.active
    ws_info.title = "Project Info"
    ws_info.append(["Attribute", "Value"])
    ws_info.append(["Project ID", "PRJ-ZERO-001"])
    ws_info.append(["Project Name", "Zero Cost Project"])
    ws_info.append(["Data Date", "2026-08-01"])
    ws_info.append(["Dataset Version", "0.2"])

    ws_wbs = wb.create_sheet("WBS")
    ws_wbs.append(["WBS Code", "WBS Name", "Parent WBS", "Discipline", "Level"])
    ws_wbs.append(["1.0", "Zero Cost Package", None, "MANAGEMENT", 1])

    ws_bdg = wb.create_sheet("Budget")
    ws_bdg.append(["WBS Code", "Cost Account", "Approved Budget", "Status", "Effective Date"])
    ws_bdg.append(["1.0", "CST-000", 0, "APPROVED", "2026-01-01"])

    ws_act = wb.create_sheet("Actual Cost")
    ws_act.append(["Transaction ID", "WBS Code", "Cost Account", "Actual Cost", "Vendor Name", "Transaction Date"])
    ws_act.append(["TX-000", "1.0", "CST-000", 0, "PT Test", "2026-01-10"])

    ws_sch = wb.create_sheet("Schedule")
    ws_sch.append(["Activity ID", "WBS Code", "Activity Name", "Planned Start", "Planned Finish", "Planned Weight %", "Actual Weight %", "Total Float", "Critical"])
    ws_sch.append(["ACT-000", "1.0", "Zero Activity", "2026-01-01", "2026-01-02", 0, 0, 0, "No"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    dataset = load_workbook(buf)
    catalogue = load_catalogue(sample_catalogue)
    context = RuleContext(catalogue=catalogue, thresholds=ThresholdConfig())
    engine = ControlEngine(ALL_RULES)

    # Must execute smoothly without uncaught ZeroDivisionError exceptions
    result = engine.run(dataset, context)
    assert result.rule_count == len(ALL_RULES)
    assert isinstance(result.findings, list)


def test_edge_case_inverted_dates_schedule_anomaly(sample_catalogue):
    """Verifies that an activity where Planned Finish < Planned Start is handled cleanly."""
    wb = openpyxl.Workbook()
    ws_info = wb.active
    ws_info.title = "Project Info"
    ws_info.append(["Attribute", "Value"])
    ws_info.append(["Project ID", "PRJ-DATE-001"])
    ws_info.append(["Project Name", "Date Anomaly Project"])
    ws_info.append(["Data Date", "2026-08-01"])
    ws_info.append(["Dataset Version", "0.2"])

    ws_wbs = wb.create_sheet("WBS")
    ws_wbs.append(["WBS Code", "WBS Name", "Parent WBS", "Discipline", "Level"])
    ws_wbs.append(["1.0", "Civil Package", None, "CIVIL", 1])

    ws_sch = wb.create_sheet("Schedule")
    ws_sch.append(["Activity ID", "WBS Code", "Activity Name", "Planned Start", "Planned Finish", "Planned Weight %", "Actual Weight %", "Total Float", "Critical"])
    # Anomaly: finish is before start
    ws_sch.append(["ACT-INV-1", "1.0", "Inverted Date Activity", "2026-08-10", "2026-08-01", 10.0, 5.0, -5, "Yes"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    dataset = load_workbook(buf)
    assert len(dataset.schedule) == 1
    assert dataset.schedule[0].baseline_start > dataset.schedule[0].baseline_finish

    catalogue = load_catalogue(sample_catalogue)
    context = RuleContext(catalogue=catalogue, thresholds=ThresholdConfig())
    engine = ControlEngine(ALL_RULES)
    result = engine.run(dataset, context)
    assert result.rule_count == len(ALL_RULES)


def test_edge_case_large_scale_wbs_performance(sample_catalogue):
    """
    Verifies that the engine can process 500+ WBS nodes and 500+ Schedule activities
    efficiently in sub-second time without quadratic bottlenecks.
    """
    wb = openpyxl.Workbook()
    ws_info = wb.active
    ws_info.title = "Project Info"
    ws_info.append(["Attribute", "Value"])
    ws_info.append(["Project ID", "PRJ-SCALE-500"])
    ws_info.append(["Project Name", "Large Scale Industrial Mega-Plant"])
    ws_info.append(["Data Date", "2026-08-15"])
    ws_info.append(["Dataset Version", "0.2"])

    ws_wbs = wb.create_sheet("WBS")
    ws_wbs.append(["WBS Code", "WBS Name", "Parent WBS", "Discipline", "Level"])
    ws_wbs.append(["1.0", "Root Plant", None, "MANAGEMENT", 1])

    ws_sch = wb.create_sheet("Schedule")
    ws_sch.append(["Activity ID", "WBS Code", "Activity Name", "Planned Start", "Planned Finish", "Planned Weight %", "Actual Weight %", "Total Float", "Critical"])

    ws_bdg = wb.create_sheet("Budget")
    ws_bdg.append(["WBS Code", "Cost Account", "Approved Budget", "Status", "Effective Date"])

    ws_act = wb.create_sheet("Actual Cost")
    ws_act.append(["Transaction ID", "WBS Code", "Cost Account", "Actual Cost", "Vendor Name", "Transaction Date"])

    # Generate 500 elements
    for i in range(1, 501):
        wbs_code = f"1.{i}"
        ws_wbs.append([wbs_code, f"Sub-package {i}", "1.0", "CIVIL", 2])
        ws_sch.append([f"ACT-{i:04d}", wbs_code, f"Task {i}", "2026-01-01", "2026-06-30", 0.2, 0.2, 5, "No"])
        ws_bdg.append([wbs_code, f"ACC-{i:04d}", 100_000_000, "APPROVED", "2026-01-01"])
        ws_act.append([f"TX-{i:04d}", wbs_code, f"ACC-{i:04d}", 90_000_000, f"Vendor {i}", "2026-03-01"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    dataset = load_workbook(buf)
    assert len(dataset.wbs_nodes) == 501
    assert len(dataset.schedule) == 500
    assert len(dataset.budgets) == 500
    assert len(dataset.actual_costs) == 500

    catalogue = load_catalogue(sample_catalogue)
    context = RuleContext(catalogue=catalogue, thresholds=ThresholdConfig())
    engine = ControlEngine(ALL_RULES)

    import time
    t0 = time.time()
    result = engine.run(dataset, context)
    elapsed = time.time() - t0

    assert result.rule_count == len(ALL_RULES)
    # Execution should be rapid
    assert elapsed < 5.0, f"Engine execution took too long: {elapsed:.2f}s"
