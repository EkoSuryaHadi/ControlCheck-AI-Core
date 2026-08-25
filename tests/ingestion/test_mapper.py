from __future__ import annotations

from dataclasses import replace
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from controlcheck.ingestion.extractor import extract_workbook
from controlcheck.ingestion.mapper import (
    DomainStatus,
    IssueSeverity,
    map_extracted_workbook,
)
from controlcheck.ingestion.profile import load_mapping_profile
from controlcheck.loader import load_workbook
from controlcheck.models import ProgressRecord, ProjectInfo, ScheduleActivity


@pytest.fixture()
def mapping_profile(project_root: Path):
    return load_mapping_profile(project_root / "data/controlcheck_mapping_profile_v0.1.json")


@pytest.fixture()
def golden_bytes(project_root: Path) -> bytes:
    return (project_root / "data/ControlCheck_AI_Golden_Positive_Dataset_v0.2.xlsx").read_bytes()


@pytest.fixture()
def invalid_progress_workbook(golden_bytes: bytes) -> bytes:
    book = openpyxl.load_workbook(BytesIO(golden_bytes))
    book["Progress"].append(
        ["PRG-INVALID", "2026-08-31", "1.0", "25%", "not-a-percent", "0%", "Open"]
    )
    payload = BytesIO()
    book.save(payload)
    book.close()
    return payload.getvalue()


def _edited_workbook(data: bytes, edit) -> bytes:
    book = openpyxl.load_workbook(BytesIO(data))
    edit(book)
    payload = BytesIO()
    book.save(payload)
    book.close()
    return payload.getvalue()


def test_golden_mapping_has_exact_canonical_counts(golden_bytes, mapping_profile):
    mapped = map_extracted_workbook(
        extract_workbook(golden_bytes, mapping_profile), mapping_profile
    )

    assert {
        name: len([row for row in rows if row.record is not None])
        for name, rows in mapped.rows_by_domain.items()
    } == {
        "wbs": 12,
        "budget": 9,
        "actual_cost": 73,
        "commitments": 6,
        "schedule": 13,
        "progress": 36,
    }
    assert set(mapped.domain_statuses.values()) == {DomainStatus.valid}
    assert mapped.error_count == 0
    assert mapped.warning_count == 0
    assert mapped.project == ProjectInfo(
        project_id="PRJ-CCAI-001",
        project_name="EPC Gas Compression Facility Expansion",
    )


def test_golden_mapping_exactly_matches_legacy_models_in_all_six_domains(
    golden_bytes, mapping_profile
):
    mapped = map_extracted_workbook(
        extract_workbook(golden_bytes, mapping_profile), mapping_profile
    )
    legacy = load_workbook(BytesIO(golden_bytes))
    expected = {
        "wbs": legacy.wbs_nodes,
        "budget": legacy.budgets,
        "actual_cost": legacy.actual_costs,
        "commitments": legacy.commitments,
        "schedule": legacy.schedule,
        "progress": legacy.progress,
    }

    for domain, legacy_records in expected.items():
        canonical = [
            row.record for row in mapped.rows_by_domain[domain] if row.record is not None
        ]
        assert [record.model_dump(mode="json") for record in canonical] == [
            record.model_dump(mode="json") for record in legacy_records
        ]


def test_golden_a9990_actual_date_anomaly_remains_canonical(
    golden_bytes, mapping_profile
):
    mapped = map_extracted_workbook(
        extract_workbook(golden_bytes, mapping_profile), mapping_profile
    )

    a9990 = next(
        row
        for row in mapped.rows_by_domain["schedule"]
        if row.record is not None and row.record.activity_id == "A9990"
    )
    assert a9990.record.actual_finish < a9990.record.actual_start
    assert a9990.issues == ()


def test_boundary_workbook_maps_every_governed_row(project_root, mapping_profile):
    boundary = (
        project_root / "data/ControlCheck_AI_Boundary_Negative_Dataset_v0.2.xlsx"
    ).read_bytes()

    mapped = map_extracted_workbook(extract_workbook(boundary, mapping_profile), mapping_profile)

    assert {
        domain: sum(row.record is not None for row in rows)
        for domain, rows in mapped.rows_by_domain.items()
    } == {
        "wbs": 1,
        "budget": 1,
        "actual_cost": 1,
        "commitments": 1,
        "schedule": 1,
        "progress": 2,
    }
    assert set(mapped.domain_statuses.values()) == {DomainStatus.valid}


def test_mapping_normalizes_values_and_preserves_exact_source_lineage(
    golden_bytes, mapping_profile
):
    def edit(book):
        schedule = book["Schedule"]
        schedule.cell(row=4, column=1, value="  A1000  ")
        schedule.cell(row=4, column=9, value="25%")
        schedule.cell(row=4, column=11, value="4.0")
        schedule.cell(row=4, column=12, value="yes")

    workbook = _edited_workbook(golden_bytes, edit)
    mapped = map_extracted_workbook(extract_workbook(workbook, mapping_profile), mapping_profile)

    result = mapped.rows_by_domain["schedule"][0]
    assert isinstance(result.record, ScheduleActivity)
    assert result.record.activity_id == "A1000"
    assert result.record.planned_progress == pytest.approx(0.25)
    assert result.record.total_float_days == 4
    assert result.record.critical is True
    assert result.record.source.sheet == "Schedule"
    assert result.record.source.row_number == 4
    assert [issue.code for issue in result.issues] == [
        "normalized_string",
        "coerced_percentage",
        "coerced_integer",
        "coerced_boolean",
    ]
    assert all(issue.severity is IssueSeverity.warning for issue in result.issues)
    assert mapped.domain_statuses["schedule"] is DomainStatus.warning


@pytest.mark.parametrize(
    ("sheet_name", "column", "value", "domain", "code"),
    [
        ("Progress", 2, "not-a-date", "progress", "invalid_date"),
        ("Progress", 5, "not-a-number", "progress", "invalid_decimal"),
        ("Schedule", 11, "1.5", "schedule", "invalid_integer"),
        ("Schedule", 12, "sometimes", "schedule", "invalid_boolean"),
    ],
)
def test_impossible_scalar_conversion_is_raw_only_with_stable_code(
    golden_bytes, mapping_profile, sheet_name, column, value, domain, code
):
    workbook = _edited_workbook(
        golden_bytes,
        lambda book: book[sheet_name].cell(row=4, column=column, value=value),
    )

    mapped = map_extracted_workbook(extract_workbook(workbook, mapping_profile), mapping_profile)
    result = mapped.rows_by_domain[domain][0]

    assert result.record is None
    assert result.issues[0].code == code
    assert result.issues[0].severity is IssueSeverity.error
    assert mapped.domain_statuses[domain] is DomainStatus.blocked


def test_missing_required_value_is_raw_only(golden_bytes, mapping_profile):
    workbook = _edited_workbook(
        golden_bytes,
        lambda book: book["Budget"].cell(row=4, column=1, value="   "),
    )

    mapped = map_extracted_workbook(extract_workbook(workbook, mapping_profile), mapping_profile)
    result = mapped.rows_by_domain["budget"][0]

    assert result.record is None
    assert [(issue.code, issue.field) for issue in result.issues] == [
        ("missing_required_value", "budget_id")
    ]
    assert mapped.domain_statuses["budget"] is DomainStatus.blocked
    assert mapped.domain_statuses["actual_cost"] is DomainStatus.valid


def test_structurally_contradictory_baseline_dates_are_raw_only(
    golden_bytes, mapping_profile
):
    def edit(book):
        schedule = book["Schedule"]
        schedule.cell(row=4, column=5, value="2026-04-01")
        schedule.cell(row=4, column=6, value="2026-03-01")

    workbook = _edited_workbook(golden_bytes, edit)
    mapped = map_extracted_workbook(extract_workbook(workbook, mapping_profile), mapping_profile)
    result = mapped.rows_by_domain["schedule"][0]

    assert result.record is None
    assert [(issue.code, issue.field) for issue in result.issues] == [
        ("contradictory_dates", "baseline_finish")
    ]
    assert mapped.domain_statuses["schedule"] is DomainStatus.blocked


def test_duplicate_wbs_identity_blocks_every_wbs_dependent_domain(
    golden_bytes, mapping_profile
):
    def edit(book):
        wbs = book["WBS"]
        wbs.cell(row=5, column=1, value="1.0")
        wbs.cell(row=5, column=3).value = None

    workbook = _edited_workbook(
        golden_bytes,
        edit,
    )

    mapped = map_extracted_workbook(extract_workbook(workbook, mapping_profile), mapping_profile)

    duplicate = mapped.rows_by_domain["wbs"][1]
    assert duplicate.record is None
    assert duplicate.issues[0].code == "duplicate_wbs_code"
    assert set(mapped.domain_statuses.values()) == {DomainStatus.blocked}


def test_wbs_cycle_and_its_descendants_are_all_raw_only(golden_bytes, mapping_profile):
    def edit(book):
        wbs = book["WBS"]
        wbs.cell(row=4, column=3, value="1.1")

    workbook = _edited_workbook(golden_bytes, edit)
    mapped = map_extracted_workbook(extract_workbook(workbook, mapping_profile), mapping_profile)
    by_code = {
        extracted.values["wbs_code"]: result
        for extracted, result in zip(
            extract_workbook(workbook, mapping_profile).rows_by_domain["wbs"],
            mapped.rows_by_domain["wbs"],
        )
    }

    assert by_code["1.0"].record is None
    assert by_code["1.1"].record is None
    assert by_code["1.2"].record is None
    assert by_code["1.0"].issues[0].code == "cyclic_wbs_parent"
    assert by_code["1.1"].issues[0].code == "cyclic_wbs_parent"
    assert by_code["1.2"].issues[0].code == "invalid_wbs_parent"
    assert set(mapped.domain_statuses.values()) == {DomainStatus.blocked}


def test_wbs_invalid_parent_removal_propagates_transitively(
    golden_bytes, mapping_profile
):
    def edit(book):
        wbs = book["WBS"]
        wbs.cell(row=4, column=3, value="missing-root")
        wbs.cell(row=5, column=3, value="1.0")
        wbs.cell(row=6, column=3, value="1.1")

    workbook = _edited_workbook(golden_bytes, edit)
    extracted = extract_workbook(workbook, mapping_profile)
    mapped = map_extracted_workbook(extracted, mapping_profile)
    by_code = {
        source.values["wbs_code"]: result
        for source, result in zip(
            extracted.rows_by_domain["wbs"], mapped.rows_by_domain["wbs"]
        )
    }

    assert all(by_code[code].record is None for code in ("1.0", "1.1", "1.2"))
    assert all(
        by_code[code].issues[0].code == "invalid_wbs_parent"
        for code in ("1.0", "1.1", "1.2")
    )
    canonical_codes = {
        row.record.wbs_code
        for row in mapped.rows_by_domain["wbs"]
        if row.record is not None
    }
    assert all(
        row.record.parent_wbs is None or row.record.parent_wbs in canonical_codes
        for row in mapped.rows_by_domain["wbs"]
        if row.record is not None
    )


def test_empty_wbs_master_blocks_all_domains_with_stable_domain_issue(
    golden_bytes, mapping_profile
):
    def edit(book):
        wbs = book["WBS"]
        wbs.delete_rows(4, wbs.max_row - 3)

    workbook = _edited_workbook(golden_bytes, edit)
    mapped = map_extracted_workbook(extract_workbook(workbook, mapping_profile), mapping_profile)

    assert mapped.rows_by_domain["wbs"] == []
    assert set(mapped.domain_statuses.values()) == {DomainStatus.blocked}
    assert mapped.domain_issues["wbs"][0].code == "empty_wbs_master"
    assert mapped.error_count == 1


def test_missing_fact_sheet_blocks_only_that_domain(golden_bytes, mapping_profile):
    def edit(book):
        del book["Commitments"]

    workbook = _edited_workbook(golden_bytes, edit)
    mapped = map_extracted_workbook(extract_workbook(workbook, mapping_profile), mapping_profile)

    assert mapped.domain_statuses["commitments"] is DomainStatus.blocked
    assert all(
        status is DomainStatus.valid
        for domain, status in mapped.domain_statuses.items()
        if domain != "commitments"
    )
    assert mapped.error_count == 1


def test_missing_wbs_sheet_blocks_all_domains(golden_bytes, mapping_profile):
    def edit(book):
        del book["WBS"]

    workbook = _edited_workbook(golden_bytes, edit)
    mapped = map_extracted_workbook(extract_workbook(workbook, mapping_profile), mapping_profile)

    assert set(mapped.domain_statuses.values()) == {DomainStatus.blocked}
    assert mapped.error_count == 1


def test_blank_project_metadata_is_not_stringified(golden_bytes, mapping_profile):
    extracted = extract_workbook(golden_bytes, mapping_profile)
    extracted = replace(
        extracted,
        project_values={**extracted.project_values, "project_id": None, "project_name": "   "},
    )

    mapped = map_extracted_workbook(extracted, mapping_profile)

    assert mapped.project is None
    assert [(issue.code, issue.field) for issue in mapped.project_issues] == [
        ("missing_project_value", "project_id"),
        ("missing_project_value", "project_name"),
    ]
    assert mapped.error_count == 2


def test_percent_syntax_is_invalid_for_non_percentage_decimal(
    golden_bytes, mapping_profile
):
    workbook = _edited_workbook(
        golden_bytes,
        lambda book: book["Budget"].cell(row=4, column=5, value="10%"),
    )

    mapped = map_extracted_workbook(extract_workbook(workbook, mapping_profile), mapping_profile)
    result = mapped.rows_by_domain["budget"][0]

    assert result.record is None
    assert [(issue.code, issue.field) for issue in result.issues] == [
        ("invalid_decimal", "budget_amount")
    ]


@pytest.mark.parametrize("numeric_boundary", [0, 0.5])
def test_numeric_date_boundary_has_invalid_date_reason_code(
    golden_bytes, mapping_profile, numeric_boundary
):
    extracted = extract_workbook(golden_bytes, mapping_profile)
    budgets = list(extracted.rows_by_domain["budget"])
    budgets[0] = replace(
        budgets[0],
        values={**budgets[0].values, "effective_date": numeric_boundary},
    )
    extracted = replace(
        extracted,
        rows_by_domain={**extracted.rows_by_domain, "budget": budgets},
    )

    mapped = map_extracted_workbook(extracted, mapping_profile)
    result = mapped.rows_by_domain["budget"][0]

    assert result.record is None
    assert result.issues[0].code == "invalid_date"
    assert result.issues[0].field == "effective_date"


def test_duplicate_source_key_rejects_later_row_deterministically(
    golden_bytes, mapping_profile
):
    extracted = extract_workbook(golden_bytes, mapping_profile)
    progress = list(extracted.rows_by_domain["progress"])
    progress[1] = replace(progress[1], source_key=progress[0].source_key)
    extracted = replace(
        extracted,
        rows_by_domain={**extracted.rows_by_domain, "progress": progress},
    )

    first = map_extracted_workbook(extracted, mapping_profile)
    second = map_extracted_workbook(extracted, mapping_profile)

    assert first.rows_by_domain["progress"][0].record is not None
    duplicate = first.rows_by_domain["progress"][1]
    assert duplicate.record is None
    assert duplicate.issues[0].code == "duplicate_source_key"
    assert first.rows_by_domain["progress"] == second.rows_by_domain["progress"]
    assert first.domain_statuses["progress"] is DomainStatus.blocked


def test_orphan_fact_wbs_stays_canonical_for_data_quality_rules(
    golden_bytes, mapping_profile
):
    mapped = map_extracted_workbook(
        extract_workbook(golden_bytes, mapping_profile), mapping_profile
    )

    orphan = next(
        row
        for row in mapped.rows_by_domain["actual_cost"]
        if row.record is not None and row.record.transaction_id == "ACT-9004"
    )
    assert orphan.record.wbs_code == "9.9"
    assert orphan.issues == ()
    assert mapped.domain_statuses["actual_cost"] is DomainStatus.valid


def test_multiple_issues_have_deterministic_profile_field_order(
    golden_bytes, mapping_profile
):
    def edit(book):
        progress = book["Progress"]
        progress.cell(row=4, column=1, value=" ")
        progress.cell(row=4, column=2, value="bad-date")
        progress.cell(row=4, column=4, value="bad-planned")
        progress.cell(row=4, column=5, value="bad-actual")

    workbook = _edited_workbook(golden_bytes, edit)
    extracted = extract_workbook(workbook, mapping_profile)

    first = map_extracted_workbook(extracted, mapping_profile)
    second = map_extracted_workbook(extracted, mapping_profile)

    assert first.rows_by_domain["progress"][0].issues == second.rows_by_domain["progress"][0].issues
    assert [issue.field for issue in first.rows_by_domain["progress"][0].issues] == [
        "progress_id",
        "period",
        "planned_progress",
        "actual_progress",
    ]


def test_hard_invalid_progress_is_raw_only_and_blocks_progress(
    invalid_progress_workbook, mapping_profile
):
    mapped = map_extracted_workbook(
        extract_workbook(invalid_progress_workbook, mapping_profile), mapping_profile
    )

    result = mapped.rows_by_domain["progress"][-1]
    assert result.record is None
    assert result.issues[0].code == "invalid_decimal"
    assert mapped.domain_statuses["progress"] is DomainStatus.blocked
    assert mapped.error_count == 1
    assert mapped.warning_count == 2
