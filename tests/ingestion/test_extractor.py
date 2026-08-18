from __future__ import annotations

from decimal import Decimal
from hashlib import sha256
from io import BytesIO
import json
from pathlib import Path

import openpyxl
import pytest

from controlcheck.ingestion import extractor
from controlcheck.ingestion.extractor import _json_safe, extract_workbook
from controlcheck.ingestion.profile import load_mapping_profile


@pytest.fixture()
def mapping_profile(project_root: Path):
    return load_mapping_profile(project_root / "data/controlcheck_mapping_profile_v0.1.json")


@pytest.fixture()
def golden_bytes(project_root: Path) -> bytes:
    return (project_root / "data/ControlCheck_AI_Golden_Positive_Dataset_v0.2.xlsx").read_bytes()


@pytest.fixture()
def invalid_progress_workbook(golden_bytes: bytes) -> bytes:
    book = openpyxl.load_workbook(BytesIO(golden_bytes))
    sheet = book["Progress"]
    sheet.append(["PRG-INVALID", "2026-08-31", "WBS-001", "25%", "not-a-percent", "0%", "Open"])
    result = BytesIO()
    book.save(result)
    book.close()
    return result.getvalue()


def test_golden_extractor_retains_all_domain_rows(golden_bytes, mapping_profile):
    extracted = extract_workbook(golden_bytes, mapping_profile)

    assert {name: len(rows) for name, rows in extracted.rows_by_domain.items()} == {
        "wbs": 12,
        "budget": 9,
        "actual_cost": 73,
        "commitments": 6,
        "schedule": 13,
        "progress": 36,
    }
    assert extracted.template_errors == []
    assert extracted.project_values == {
        "project_id": "PRJ-CCAI-001",
        "project_name": "EPC Gas Compression Facility Expansion",
        "dataset_version": "0.2",
        "catalogue_version": "0.2",
        "data_date": "2026-08-15T00:00:00",
        "fixture_type": "golden_positive",
        "purpose": "Controlled exhaustive positive validation.",
    }
    assert extracted.workbook_sha256 == sha256(golden_bytes).hexdigest()


def test_non_empty_invalid_row_is_retained(invalid_progress_workbook, mapping_profile):
    extracted = extract_workbook(invalid_progress_workbook, mapping_profile)

    row = extracted.rows_by_domain["progress"][-1]
    assert row.values["actual_progress"] == "not-a-percent"
    assert row.source_row_number > 1


def test_missing_governed_sheet_becomes_a_stable_template_issue(golden_bytes, mapping_profile):
    book = openpyxl.load_workbook(BytesIO(golden_bytes))
    del book["Progress"]
    payload = BytesIO()
    book.save(payload)
    book.close()

    extracted = extract_workbook(payload.getvalue(), mapping_profile)

    assert extracted.rows_by_domain["progress"] == []
    assert [(issue.code, issue.domain, issue.sheet_name) for issue in extracted.template_errors] == [
        ("missing_sheet", "progress", "Progress"),
    ]


def test_malformed_headers_are_reported_in_a_stable_order_without_dropping_rows(
    golden_bytes, mapping_profile
):
    book = openpyxl.load_workbook(BytesIO(golden_bytes))
    sheet = book["Progress"]
    sheet.cell(row=3, column=2, value="progress_id")
    sheet.cell(row=3, column=8, value="unexpected")
    payload = BytesIO()
    book.save(payload)
    book.close()

    first = extract_workbook(payload.getvalue(), mapping_profile)
    second = extract_workbook(payload.getvalue(), mapping_profile)

    assert len(first.rows_by_domain["progress"]) == 36
    assert [(issue.code, issue.domain, issue.sheet_name) for issue in first.template_errors] == [
        ("duplicate_header", "progress", "Progress"),
        ("missing_required_column", "progress", "Progress"),
        ("unexpected_header", "progress", "Progress"),
    ]
    assert first.template_errors == second.template_errors


def test_extracted_values_are_json_safe_and_source_keys_are_deterministic(
    golden_bytes, mapping_profile
):
    first = extract_workbook(golden_bytes, mapping_profile)
    second = extract_workbook(golden_bytes, mapping_profile)

    first_rows = first.rows_by_domain["schedule"]
    second_rows = second.rows_by_domain["schedule"]
    assert json.loads(json.dumps(first_rows[0].values)) == first_rows[0].values
    assert [row.source_key for row in first_rows] == [row.source_key for row in second_rows]
    assert len({row.source_key for row in first_rows}) == len(first_rows)


def test_collision_prone_headers_preserve_every_cell_and_source_key_tracks_each_value(
    mapping_profile,
):
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Progress"
    sheet.append(["title"])
    sheet.append(["metadata"])
    sheet.append(["x", "x", "x__duplicate_2", None, "__unnamed_column_4"])
    sheet.append(["first", "second", "third", "fourth", "fifth"])
    payload = BytesIO()
    book.save(payload)
    book.close()

    first = extract_workbook(payload.getvalue(), mapping_profile).rows_by_domain["progress"][0]
    changed_book = openpyxl.load_workbook(BytesIO(payload.getvalue()))
    changed_book["Progress"].cell(row=4, column=4, value="changed-fourth")
    changed_payload = BytesIO()
    changed_book.save(changed_payload)
    changed_book.close()
    changed = extract_workbook(changed_payload.getvalue(), mapping_profile).rows_by_domain["progress"][0]

    assert set(first.values.values()) == {"first", "second", "third", "fourth", "fifth"}
    assert len(first.values) == 5
    assert first.source_key != changed.source_key


def test_blank_data_rows_and_pre_header_title_rows_are_excluded(golden_bytes, mapping_profile):
    book = openpyxl.load_workbook(BytesIO(golden_bytes))
    blank_row = book["Progress"].max_row + 1
    book["Progress"].cell(row=blank_row, column=1)
    payload = BytesIO()
    book.save(payload)
    book.close()

    extracted = extract_workbook(payload.getvalue(), mapping_profile)

    assert len(extracted.rows_by_domain["progress"]) == 36
    assert all(row.source_row_number != blank_row for row in extracted.rows_by_domain["progress"])
    assert "CONTROLCHECK AI — VALIDATION DATASET" not in extracted.project_values


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (Decimal("12.50"), "12.50"),
        (b"\x00\xff", "AP8="),
        (float("inf"), "inf"),
        (float("nan"), "nan"),
    ],
)
def test_json_safe_conversion_handles_non_excel_scalar_values(value, expected):
    assert _json_safe(value) == expected


def test_workbook_is_closed_when_extraction_raises(golden_bytes, mapping_profile, monkeypatch):
    book = openpyxl.load_workbook(BytesIO(golden_bytes), data_only=True, read_only=True)
    monkeypatch.setattr(extractor.openpyxl, "load_workbook", lambda *args, **kwargs: book)

    def raise_header_issue(*args, **kwargs):
        raise RuntimeError("forced extraction failure")

    monkeypatch.setattr(extractor, "_header_issues", raise_header_issue)

    with pytest.raises(RuntimeError, match="forced extraction failure"):
        extract_workbook(golden_bytes, mapping_profile)

    assert book._archive.fp is None
