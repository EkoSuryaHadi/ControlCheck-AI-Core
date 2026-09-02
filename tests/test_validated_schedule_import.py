from datetime import date, datetime
from io import BytesIO

import openpyxl

from controlcheck.ingestion.validated_import import build_schedule_workbook


def test_build_schedule_workbook_maps_ms_project_tasks_into_governed_schedule():
    source = openpyxl.Workbook()
    tasks = source.active
    tasks.title = "Tasks"
    tasks.append([
        "ID", "Name", "Outline Number", "Baseline Start", "Baseline Finish",
        "Start", "Finish", "% Complete", "Total Slack",
    ])
    tasks.append(["10", "Install compressor", "1.2", datetime(2026, 1, 1), datetime(2026, 1, 10), datetime(2026, 1, 2), None, 35, -2])
    buffer = BytesIO()
    source.save(buffer)

    payload = build_schedule_workbook(
        buffer.getvalue(),
        "tasks.xlsx",
        project_code="ABACUS-1",
        project_name="ABD",
        data_date=date(2026, 1, 15),
        preset="msproject",
    )

    workbook = openpyxl.load_workbook(BytesIO(payload), data_only=True)
    assert {"Project_Info", "Schedule"}.issubset(workbook.sheetnames)
    schedule = workbook["Schedule"]
    headers = [cell.value for cell in schedule[3]]
    values = dict(zip(headers, [cell.value for cell in schedule[4]]))
    assert values["activity_id"] == "10"
    assert values["activity_name"] == "Install compressor"
    assert values["wbs_code"] == "1.2"
    assert values["actual_progress"] == 35
    assert values["total_float_days"] == -2
    assert values["critical"] is True


def test_build_schedule_workbook_uses_snapshot_dataset_version_that_fits_database_limit():
    source = openpyxl.Workbook()
    tasks = source.active
    tasks.title = "Tasks"
    tasks.append([
        "ID", "Name", "Outline Number", "Baseline Start", "Baseline Finish",
        "Start", "Finish", "% Complete", "Total Slack",
    ])
    tasks.append(["10", "Install compressor", "1.2", "2026-01-01", "2026-01-10", None, None, 0, 0])
    buffer = BytesIO()
    source.save(buffer)

    payload = build_schedule_workbook(
        buffer.getvalue(),
        "tasks.xlsx",
        project_code="ABACUS-1",
        project_name="ABD",
        data_date=date(2026, 1, 15),
        preset="msproject",
    )

    workbook = openpyxl.load_workbook(BytesIO(payload), data_only=True)
    project_values = {
        row[0].value: row[1].value
        for row in workbook["Project_Info"].iter_rows(min_row=2, max_col=2)
    }
    assert project_values["dataset_version"] == "validated-schedule-1"
    assert len(project_values["dataset_version"]) <= 20

def test_available_health_categories_only_includes_executed_rule_domains():
    from controlcheck.application import available_health_categories

    assert available_health_categories(["SCH-001", "SCH-005"]) == {"SCHEDULE"}
