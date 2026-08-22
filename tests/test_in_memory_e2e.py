from io import BytesIO
from datetime import date
from decimal import Decimal
from uuid import uuid4
import openpyxl

from controlcheck.loader import load_workbook
from controlcheck.config import load_catalogue, ThresholdConfig
from controlcheck.engine import ControlEngine, RuleContext
from controlcheck.rules import ALL_RULES
from controlcheck.health.scoring import compute_health_score
from controlcheck.ai.assistant import ProjectAIAssistant
from controlcheck.ai.safety import sanitize_ai_response, AI_SAFETY_RULES


def _build_full_lifecycle_workbook(project_id: str = "PRJ-E2E-100") -> BytesIO:
    """Build a complete Excel workbook with multi-dimensional project control data."""
    wb = openpyxl.Workbook()

    # 1. Project Info
    ws_info = wb.active
    ws_info.title = "Project Info"
    ws_info.append(["Attribute", "Value"])
    ws_info.append(["Project ID", project_id])
    ws_info.append(["Project Name", "End-to-End Modern Hospital Construction"])
    ws_info.append(["Data Date", "2026-08-15"])
    ws_info.append(["Currency", "IDR"])
    ws_info.append(["Dataset Version", "0.2"])

    # 2. WBS
    ws_wbs = wb.create_sheet("WBS")
    ws_wbs.append(["WBS Code", "WBS Name", "Parent WBS", "Discipline", "Level"])
    ws_wbs.append(["1.0", "Main Hospital Facility", None, "MANAGEMENT", 1])
    ws_wbs.append(["1.1", "Civil & Structural", "1.0", "CIVIL", 2])
    ws_wbs.append(["1.2", "MEP Installation", "1.0", "MEP", 2])

    # 3. Budget
    ws_bdg = wb.create_sheet("Budget")
    ws_bdg.append(["WBS Code", "Cost Account", "Approved Budget", "Status", "Effective Date"])
    ws_bdg.append(["1.1", "CST-CIV-01", 5_000_000_000, "APPROVED", "2026-01-01"])
    ws_bdg.append(["1.2", "CST-MEP-01", 3_000_000_000, "APPROVED", "2026-01-01"])

    # 4. Actual Cost
    ws_act = wb.create_sheet("Actual Cost")
    ws_act.append(["Transaction ID", "WBS Code", "Cost Account", "Actual Cost", "Vendor Name", "Transaction Date"])
    ws_act.append(["TX-CIV-001", "1.1", "CST-CIV-01", 5_800_000_000, "PT Beton Prima", "2026-07-20"])
    ws_act.append(["TX-MEP-001", "1.2", "CST-MEP-01", 1_200_000_000, "PT Elektrika Jaya", "2026-08-01"])

    # 5. Schedule
    ws_sch = wb.create_sheet("Schedule")
    ws_sch.append(["Activity ID", "WBS Code", "Activity Name", "Planned Start", "Planned Finish", "Planned Weight %", "Actual Weight %", "Total Float", "Critical"])
    ws_sch.append(["ACT-CIV-01", "1.1", "Substructure Piling", "2026-01-15", "2026-06-30", 50.0, 50.0, 0, "Yes"])
    ws_sch.append(["ACT-CIV-02", "1.1", "Superstructure Framing", "2026-07-01", "2026-11-30", 30.0, 10.0, -10, "Yes"])
    ws_sch.append(["ACT-MEP-01", "1.2", "HVAC Ducting Rough-in", "2026-08-01", "2026-12-15", 20.0, 0.0, 15, "No"])

    # 6. Progress
    ws_prg = wb.create_sheet("Progress")
    ws_prg.append(["WBS Code", "Cutoff Date", "Planned Progress %", "Actual Progress %"])
    ws_prg.append(["1.0", "2026-08-15", 65.0, 45.0])
    ws_prg.append(["1.1", "2026-08-15", 80.0, 60.0])
    ws_prg.append(["1.2", "2026-08-15", 20.0, 0.0])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_in_memory_full_lifecycle_pipeline(sample_catalogue):
    """
    Tests the complete end-to-end lifecycle entirely in memory:
    1. Dataset Generation & Parsing
    2. Rule Engine Execution (20 rules)
    3. Finding Extraction & Categorization
    4. Health Score Multi-Pillar Computation
    5. AI Assistant Grounding & Intent Resolution
    """
    # Step 1: Ingestion & Parsing
    buf = _build_full_lifecycle_workbook("PRJ-HOSP-001")
    dataset = load_workbook(buf)

    assert dataset.project.project_id == "PRJ-HOSP-001"
    assert dataset.project.project_name == "End-to-End Modern Hospital Construction"
    assert dataset.data_date == date(2026, 8, 15)
    assert len(dataset.wbs_nodes) == 3
    assert len(dataset.budgets) == 2
    assert len(dataset.actual_costs) == 2
    assert len(dataset.schedule) == 3
    assert len(dataset.progress) == 3

    # Step 2: Rule Engine Execution
    catalogue = load_catalogue(sample_catalogue)
    context = RuleContext(catalogue=catalogue, thresholds=ThresholdConfig())
    engine = ControlEngine(ALL_RULES)
    result = engine.run(dataset, context)

    assert result.rule_count == len(ALL_RULES)
    assert result.finding_count > 0

    # Step 3: Verify Specific Findings Triggered
    # Overrun on 1.1: Actual 5.8B > Budget 5.0B -> CST-001
    cst_findings = [f for f in result.findings if f.rule_id == "CST-001"]
    assert len(cst_findings) >= 1
    assert any("1.1" in str(f.entity_id) for f in cst_findings)

    # Schedule Negative Float on ACT-CIV-02 (-10 days) -> SCH-002 or similar
    sch_findings = [f for f in result.findings if f.category == "schedule"]
    assert len(sch_findings) >= 1

    # Step 4: Health Scoring Multi-Pillar
    findings_dicts = [
        {
            "finding_id": f.finding_id,
            "rule_id": f.rule_id,
            "category": f.category,
            "severity": f.severity,
            "business_impact": f.business_impact,
            "recommendation": f.recommendation,
            "entity_id": f.entity_id,
        }
        for f in result.findings
    ]

    health = compute_health_score(findings_dicts)
    assert 0.0 <= health.overall_score <= 100.0
    assert health.score_band in ("Healthy", "Needs Attention", "At Risk", "Critical")
    assert 0.0 <= health.category_scores["COST"].score <= 100.0
    assert 0.0 <= health.category_scores["SCHEDULE"].score <= 100.0
    assert 0.0 <= health.category_scores["PROGRESS"].score <= 100.0
    assert 0.0 <= health.category_scores["DATA_QUALITY"].score <= 100.0
    assert isinstance(health.top_drivers, list)

    # Step 5: In-Memory AI Assistant Intent & Safety
    from unittest.mock import MagicMock, patch

    mock_session = MagicMock()
    org_id = uuid4()
    proj_id = uuid4()
    assistant = ProjectAIAssistant(org_id, proj_id, session=mock_session)

    # Test Health Query with grounded data
    with patch("controlcheck.ai.assistant.get_project_health") as mock_health_tool:
        mock_health_tool.return_value = {
            "overall_score": health.overall_score,
            "score_band": health.score_band,
            "cost_score": health.category_scores["COST"].score,
            "schedule_score": health.category_scores["SCHEDULE"].score,
            "progress_score": health.category_scores["PROGRESS"].score,
            "dq_score": health.category_scores["DATA_QUALITY"].score,
            "key_drivers": [
                {
                    "rule_id": d.rule_id,
                    "description": d.description,
                    "severity": d.severity,
                    "penalty": d.penalty,
                }
                for d in health.top_drivers[:3]
            ],
        }
        health_qa = assistant.ask("Bagaimana status kesehatan proyek saat ini?")
        assert f"{health.overall_score:.1f}/100" in health_qa["answer"]
        assert health.score_band in health_qa["answer"]
        assert health_qa["confidence"] == "high"

    # Test Cost Driver Query with grounded findings
    with patch("controlcheck.ai.assistant.get_top_cost_drivers") as mock_cost_tool:
        mock_cost_tool.return_value = [
            {
                "finding_id": f.finding_id,
                "rule_id": f.rule_id,
                "entity_id": f.entity_id,
                "title": f.title,
                "description": f.description,
                "business_impact": f.business_impact,
                "recommendation": f.recommendation,
            }
            for f in cst_findings[:2]
        ]
        cost_qa = assistant.ask("Apakah ada cost overrun pada anggaran?")
        assert "CST-001" in cost_qa["answer"] or "1.1" in cost_qa["answer"]
        assert isinstance(cost_qa["evidence_references"], list)

    # Test Safety Sanitizer
    sanitized = sanitize_ai_response({"answer": "Audit confirmed findings.", "key_evidence": ["TX-CIV-001"]})
    assert sanitized["answer"] == "Audit confirmed findings."
    assert sanitized["key_evidence"] == ["TX-CIV-001"]
