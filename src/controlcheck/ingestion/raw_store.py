from __future__ import annotations

from typing import Any, BinaryIO
from pathlib import Path
from uuid import UUID, uuid4
import openpyxl

from ..loader import REQUIRED_COLUMNS
from ..logging import get_logger

logger = get_logger("ingestion.raw_store")


class RawRowItem:
    def __init__(
        self,
        sheet_name: str,
        row_number: int,
        raw_data: dict[str, Any],
        id: UUID | None = None,
    ):
        self.id = id or uuid4()
        self.sheet_name = sheet_name
        self.row_number = row_number
        self.raw_data = raw_data


def extract_raw_rows(source: Path | str | BinaryIO) -> list[RawRowItem]:
    """Extracts all non-empty raw rows with headers from an Excel workbook."""
    logger.info("Extracting raw rows from workbook source")
    book = openpyxl.load_workbook(source, data_only=True, read_only=False)
    raw_rows: list[RawRowItem] = []

    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        header_row = None
        headers: list[str] = []

        required = REQUIRED_COLUMNS.get(sheet_name)
        if required:
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                candidate = [str(value).strip() if value is not None else "" for value in row]
                if required <= set(candidate):
                    header_row, headers = row_idx, candidate
                    break
        else:
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                non_empty = [c for c in row if c is not None]
                if len(non_empty) >= 2 or (len(non_empty) == 1 and row_idx == 1):
                    header_row = row_idx
                    headers = [str(cell).strip() if cell is not None else f"col_{c_idx}" for c_idx, cell in enumerate(row)]
                    break

        if header_row is None:
            continue


        for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not any(cell is not None for cell in row):
                continue
            
            # Build row dict safely
            row_dict = {}
            for col_idx, cell_value in enumerate(row):
                header = headers[col_idx] if col_idx < len(headers) else f"col_{col_idx}"
                # Serialize datetimes / decimals to string/safe types for JSONB
                if hasattr(cell_value, "isoformat"):
                    row_dict[header] = cell_value.isoformat()
                elif cell_value is not None:
                    row_dict[header] = str(cell_value)
                else:
                    row_dict[header] = None

            raw_rows.append(RawRowItem(
                sheet_name=sheet_name,
                row_number=row_idx,
                raw_data=row_dict,
            ))

    book.close()
    logger.info("Extracted %d raw rows across %d sheet(s)", len(raw_rows), len(book.sheetnames))
    return raw_rows
