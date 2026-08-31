from __future__ import annotations

import csv
import re
from dataclasses import asdict, dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any

import openpyxl


@dataclass
class ValidationIssue:
    row: int
    field: str
    severity: str
    code: str
    message: str
    value: Any = None


MS_PROJECT_ALIASES = {
    "activity_id": ["unique id", "id", "task id"],
    "activity_name": ["name", "task name"],
    "wbs": ["wbs", "outline number"],
    "baseline_start": ["baseline start", "baseline1 start"],
    "baseline_finish": ["baseline finish", "baseline1 finish"],
    "start": ["start"],
    "finish": ["finish"],
    "duration": ["duration"],
    "progress": ["% complete", "percent complete"],
    "total_float": ["total slack", "total float", "slack"],
    "predecessors": ["predecessors"],
    "resources": ["resource names", "resources"],
    "constraint_type": ["constraint type"],
    "constraint_date": ["constraint date"],
}

STANDARD_ALIASES = {
    "activity_id": ["activity_id", "activity id", "task_code", "task id", "unique id"],
    "activity_name": ["activity_name", "activity name", "task_name", "name"],
    "wbs": ["wbs_code", "wbs code", "wbs", "wbs_id"],
    "baseline_start": ["baseline_start", "baseline start", "target_start_date", "planned start"],
    "baseline_finish": ["baseline_finish", "baseline finish", "target_end_date", "planned finish"],
    "start": ["current_start", "start", "current start"],
    "finish": ["current_finish", "finish", "current finish"],
    "duration": ["duration", "baseline_duration_days"],
    "progress": ["actual_progress_pct", "% complete", "act_pct_comp", "physical_percent_complete"],
    "total_float": ["total_float_days", "total float", "total slack", "total_float_hr_cnt"],
    "predecessors": ["predecessor_ids", "predecessors"],
    "resources": ["responsible", "resource names", "resources"],
    "constraint_type": ["constraint_type", "constraint type"],
    "constraint_date": ["constraint_date", "constraint date"],
}


def _norm(value: Any) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _find_column(headers: list[str], aliases: list[str]) -> int | None:
    normalized = [_norm(h) for h in headers]
    candidates = [_norm(a) for a in aliases]
    for idx, value in enumerate(normalized):
        if value in candidates:
            return idx
    return None


def _as_number(value: Any) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", str(value).replace(",", "").replace("%", ""))
    return float(match.group(0)) if match else None


def _as_date(value: Any) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _read_rows(data: bytes, filename: str) -> tuple[str, list[str], list[list[Any]]]:
    if filename.lower().endswith(".csv"):
        rows = list(csv.reader(StringIO(data.decode("utf-8-sig", errors="replace"))))
        return ("CSV", [str(v).strip() for v in rows[0]], rows[1:]) if rows else ("CSV", [], [])
    book = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True)
    preferred = next((name for name in book.sheetnames if any(k in _norm(name) for k in ("task", "schedule", "activity", "jadwal"))), book.sheetnames[0])
    sheet = book[preferred]
    best_row, best_headers = 1, []
    for idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
        headers = [str(v).strip() if v is not None else "" for v in row]
        if len([h for h in headers if h]) > len([h for h in best_headers if h]):
            best_row, best_headers = idx, headers
    return sheet.title, best_headers, [list(r) for r in sheet.iter_rows(min_row=best_row + 1, values_only=True)]


def _relationships(value: Any) -> list[tuple[str, str, float]]:
    if value is None or str(value).strip() == "":
        return []
    output = []
    for token in re.split(r"[,;]", str(value)):
        token = token.strip()
        if not token:
            continue
        match = re.match(r"^(.+?)(FS|SS|FF|SF)?([+-]\s*\d+(?:\.\d+)?\s*[a-zA-Z]*)?$", token, re.I)
        if not match:
            output.append((token, "FS", 0.0))
            continue
        pred, rel, lag_raw = match.group(1).strip(), (match.group(2) or "FS").upper(), match.group(3) or ""
        lag = 0.0
        lm = re.search(r"([+-])\s*(\d+(?:\.\d+)?)\s*([a-zA-Z]*)", lag_raw)
        if lm:
            lag = float(lm.group(2)) * (-1 if lm.group(1) == "-" else 1)
            unit = lm.group(3).lower()
            if unit.startswith("h"):
                lag /= 8
            elif unit.startswith("w"):
                lag *= 5
        output.append((pred, rel, lag))
    return output


def validate_workbook_bytes(data: bytes, filename: str, preset: str = "standard") -> dict[str, Any]:
    if not filename.lower().endswith((".xlsx", ".xlsm", ".csv")):
        raise ValueError("Supported validation formats are .xlsx, .xlsm, and .csv")

    sheet, headers, rows = _read_rows(data, filename)
    aliases = MS_PROJECT_ALIASES if preset == "msproject" else STANDARD_ALIASES
    mapping = {key: _find_column(headers, values) for key, values in aliases.items()}
    issues: list[ValidationIssue] = []
    for field in ("activity_id", "activity_name", "baseline_start", "baseline_finish"):
        if mapping.get(field) is None:
            issues.append(ValidationIssue(1, field, "error", "missing_required_column", f"Required column for {field} was not detected."))

    def value(row: list[Any], field: str):
        idx = mapping.get(field)
        return row[idx] if idx is not None and idx < len(row) else None

    populated = [(idx, row) for idx, row in enumerate(rows, start=2) if any(v is not None and str(v).strip() for v in row)]
    ids = {str(value(row, "activity_id")).strip() for _, row in populated if value(row, "activity_id") not in (None, "")}
    successors = {aid: 0 for aid in ids}
    seen: set[str] = set()
    preview = []
    relationship_count = logic_checks = logic_failures = 0
    open_start = excessive_lag = negative_lag = excessive_float = hard_constraints = non_fs = 0
    checked_required = missing_required = 0

    for rownum, row in populated:
        aid = str(value(row, "activity_id") or "").strip()
        name = value(row, "activity_name")
        bs_raw, bf_raw = value(row, "baseline_start"), value(row, "baseline_finish")
        start_raw, finish_raw = value(row, "start"), value(row, "finish")
        duration, progress, total_float = value(row, "duration"), value(row, "progress"), value(row, "total_float")
        predecessors, resources = value(row, "predecessors"), value(row, "resources")
        constraint_type = value(row, "constraint_type")

        for field, cell in (("activity_id", aid), ("activity_name", name), ("baseline_start", bs_raw), ("baseline_finish", bf_raw)):
            checked_required += 1
            if cell in (None, ""):
                missing_required += 1
                issues.append(ValidationIssue(rownum, field, "error", "required_value_missing", f"{field} is required."))

        if aid:
            if aid in seen:
                issues.append(ValidationIssue(rownum, "activity_id", "error", "duplicate_activity_id", f"Duplicate activity ID: {aid}", aid))
            seen.add(aid)

        progress_num = _as_number(progress)
        if progress not in (None, "") and progress_num is None:
            issues.append(ValidationIssue(rownum, "progress", "warning", "invalid_progress", "Progress is not numeric.", progress))
        elif progress_num is not None and not 0 <= progress_num <= 100:
            issues.append(ValidationIssue(rownum, "progress", "error", "progress_out_of_range", "Progress must be between 0 and 100.", progress))

        float_num = _as_number(total_float)
        if total_float in (None, ""):
            issues.append(ValidationIssue(rownum, "total_float", "warning", "float_missing", "Total Slack/Float is missing."))
        elif float_num is not None and float_num > 20:
            excessive_float += 1
            issues.append(ValidationIssue(rownum, "total_float", "warning", "excessive_float", "Total Slack/Float exceeds 20 days.", total_float))

        bs, bf, cs, cf = _as_date(bs_raw), _as_date(bf_raw), _as_date(start_raw), _as_date(finish_raw)
        for field, raw, parsed, severity in (("baseline_start", bs_raw, bs, "error"), ("baseline_finish", bf_raw, bf, "error"), ("start", start_raw, cs, "warning"), ("finish", finish_raw, cf, "warning")):
            if raw not in (None, ""):
                logic_checks += 1
                if parsed is None:
                    logic_failures += 1
                    issues.append(ValidationIssue(rownum, field, severity, f"invalid_{field}", "Date value is not recognized.", raw))
        if bs and bf:
            logic_checks += 1
            if bf < bs:
                logic_failures += 1
                issues.append(ValidationIssue(rownum, "baseline_finish", "error", "baseline_finish_before_start", "Baseline Finish is earlier than Baseline Start."))
        if cs and cf:
            logic_checks += 1
            if cf < cs:
                logic_failures += 1
                issues.append(ValidationIssue(rownum, "finish", "error", "current_finish_before_start", "Finish is earlier than Start."))

        duration_num = _as_number(duration)
        if duration_num is not None:
            logic_checks += 1
            if duration_num < 0:
                logic_failures += 1
                issues.append(ValidationIssue(rownum, "duration", "error", "negative_duration", "Duration cannot be negative.", duration))
            elif duration_num == 0 and bs and bf and bs != bf:
                issues.append(ValidationIssue(rownum, "duration", "warning", "milestone_date_mismatch", "Zero-duration task should normally have the same baseline start and finish."))

        rels = _relationships(predecessors)
        relationship_count += len(rels)
        if not rels:
            open_start += 1
            issues.append(ValidationIssue(rownum, "predecessors", "warning", "open_start", "Activity has no predecessor relationship."))
        for pred, rel, lag in rels:
            logic_checks += 1
            if pred == aid:
                logic_failures += 1
                issues.append(ValidationIssue(rownum, "predecessors", "error", "self_predecessor", "Activity cannot be its own predecessor."))
            elif pred not in ids:
                logic_failures += 1
                issues.append(ValidationIssue(rownum, "predecessors", "warning", "predecessor_not_found", f"Predecessor {pred} was not found."))
            else:
                successors[pred] = successors.get(pred, 0) + 1
            if rel != "FS":
                non_fs += 1
            if lag < 0:
                negative_lag += 1
                issues.append(ValidationIssue(rownum, "predecessors", "warning", "negative_lag", f"Negative lag detected ({lag:g} days)."))
            elif lag > 10:
                excessive_lag += 1
                issues.append(ValidationIssue(rownum, "predecessors", "warning", "excessive_lag", f"Lag exceeds 10 days ({lag:g} days)."))

        if resources in (None, ""):
            issues.append(ValidationIssue(rownum, "resources", "warning", "resource_missing", "No responsible resource/owner was supplied."))
        if constraint_type not in (None, "") and _norm(constraint_type) not in {"as soon as possible", "as late as possible", "asap", "alap"}:
            hard_constraints += 1
            issues.append(ValidationIssue(rownum, "constraint_type", "warning", "hard_constraint", "Hard date constraint detected; verify justification.", constraint_type))

        if len(preview) < 10:
            preview.append({"row": rownum, "activity_id": aid, "activity_name": name, "wbs": value(row, "wbs"), "baseline_start": bs_raw, "baseline_finish": bf_raw, "start": start_raw, "finish": finish_raw, "duration": duration, "progress": progress, "total_float": total_float, "predecessors": predecessors})

    open_finish = 0
    for aid, count in successors.items():
        if count == 0:
            open_finish += 1

    error_count = sum(i.severity == "error" for i in issues)
    warning_count = sum(i.severity == "warning" for i in issues)
    completeness = 100 if checked_required == 0 else round(100 * (checked_required - missing_required) / checked_required)
    data_score = max(0, min(100, completeness - min(45, error_count * 4) - min(15, warning_count)))
    core_logic = 100 if logic_checks == 0 else round(100 * (logic_checks - logic_failures) / logic_checks)
    schedule_score = max(0, min(100, core_logic - min(25, open_start + open_finish) - min(15, excessive_lag * 2 + negative_lag * 3) - min(10, excessive_float) - min(10, hard_constraints)))

    return {
        "filename": filename,
        "preset": preset,
        "sheet": sheet,
        "total_rows": len(populated),
        "detected_headers": headers,
        "mapping": mapping,
        "preview": preview,
        "issues": [asdict(i) for i in issues[:500]],
        "error_count": error_count,
        "warning_count": warning_count,
        "data_quality_score": data_score,
        "schedule_quality_score": schedule_score,
        "logic_check_count": logic_checks,
        "logic_failure_count": logic_failures,
        "relationship_metrics": {
            "relationship_count": relationship_count,
            "logic_density": round(relationship_count / len(populated), 2) if populated else 0,
            "open_start_count": open_start,
            "open_finish_count": open_finish,
            "excessive_lag_count": excessive_lag,
            "negative_lag_count": negative_lag,
            "excessive_float_count": excessive_float,
            "hard_constraint_count": hard_constraints,
            "non_fs_relationship_count": non_fs,
        },
        "can_import": error_count == 0 and len(populated) > 0,
        "truncated_issues": len(issues) > 500,
    }
