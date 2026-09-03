"""Microsoft Project (.mpp) → ControlCheck workbook converter.

MPXJ (via its official Python/JPype bridge) parses the proprietary binary
format; we map MS Project tasks onto the standard ControlCheck workbook
shape (Schedule / WBS / Progress / Budget / Actual_Cost sheets) so the
existing mapping-profile pipeline can ingest schedules straight from MS
Project files.

The JVM is started lazily (once per process) — safe inside Celery prefork
workers, where each child converts independently.
"""

from __future__ import annotations

import logging
import os
import tempfile
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook

logger = logging.getLogger(__name__)

DEFAULT_SHEET_TITLE = "MS PROJECT SCHEDULE IMPORT"
_STATUS_COMPLETE = "Complete"
_STATUS_IN_PROGRESS = "In Progress"
_STATUS_NOT_STARTED = "Not Started"
_PROGRESS_STATUS = "Approved"


class MppConversionError(ValueError):
    """Raised when a .mpp file cannot be parsed or mapped."""


def _as_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[: len("2026-01-05T08:00")], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        logger.warning("Unparseable MPXJ date %r", value)
        return None


def _as_days(duration: Any) -> float:
    """MPXJ Duration string like ``0.0d`` or ``5.0d`` → float days."""
    if duration is None:
        return 0.0
    text = str(duration).strip()
    try:
        return float(text.rstrip("dDd "))
    except ValueError:
        return 0.0


def _cell_date(value: date | None) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%d 00:00:00")


def _progress_status(pct: float) -> str:
    if pct >= 0.995:
        return _STATUS_COMPLETE
    if pct <= 0.005:
        return _STATUS_NOT_STARTED
    return _STATUS_IN_PROGRESS


def _clamp(value: float) -> float:
    return max(0.0, min(1.0, value))


class MppConverter:
    """Converts .mpp bytes into a ControlCheck-compatible xlsx workbook."""

    def __init__(self, java_home: str | None = None) -> None:
        self.java_home = java_home
        self._jvm_ready = False

    def _ensure_jvm(self) -> None:
        if self._jvm_ready:
            return
        try:
            import jpype  # noqa: PLC0415
            import mpxj  # noqa: F401, PLC0415
        except ImportError as exc:
            raise MppConversionError(
                "MS Project (.mpp) ingestion requires the 'mpxj' and 'jpype1' "
                "packages — install them on the worker image."
            ) from exc
        if self.java_home:
            os.environ["JAVA_HOME"] = self.java_home
        if not jpype.isJVMStarted():
            jpype.startJVM()
        self._jvm_ready = True

    def _read_project(self, path: Path) -> Any:
        from org.mpxj.reader import UniversalProjectReader  # noqa: PLC0415

        try:
            return UniversalProjectReader().read(str(path))
        except Exception as exc:  # MPXJ throws java exceptions on corrupt files
            raise MppConversionError(
                f"Could not parse MS Project file: {exc}"
            ) from exc

    def to_workbook_bytes(self, data: bytes, filename: str = "upload.mpp") -> bytes:
        self._ensure_jvm()
        suffix = Path(filename).suffix.lower() or ".mpp"
        with tempfile.NamedTemporaryFile(suffix=suffix) as tmp:
            tmp.write(data)
            tmp.flush()
            project = self._read_project(Path(tmp.name))
        return self._build_workbook(project)

    # ------------------------------------------------------------------ #

    def _build_workbook(self, project: Any) -> bytes:
        props = project.getProjectProperties()
        title = str(getattr(props, "getProjectTitle", lambda: None)() or DEFAULT_SHEET_TITLE)
        status_date = _as_date(getattr(props, "getStatusDate", lambda: None)())
        project_start = _as_date(getattr(props, "getStartDate", lambda: None)())

        wbs_map: dict[str, dict[str, Any]] = {}
        schedule_rows: list[list[Any]] = []
        budget_by_wbs: dict[str, float] = {}
        actual_by_wbs: dict[str, float] = {}

        tasks = list(project.getTasks())
        if not tasks:
            raise MppConversionError("Project file contains no tasks.")

        for task in tasks:
            wbs = str(task.getWBS() or "").strip()
            name = str(task.getName() or "").strip()
            if not wbs or wbs in ("0",):
                continue  # project root pseudo-task
            level = int(task.getOutlineLevel() or 0)
            parent = None
            for getter in ("getParentTask", "getParent"):
                if hasattr(task, getter):
                    parent = getattr(task, getter)()
                    break
            parent_wbs = ""
            if parent is not None:
                pwbs = str(parent.getWBS() or "").strip()
                if pwbs and pwbs != "0":
                    parent_wbs = pwbs
            if wbs not in wbs_map:
                wbs_map[wbs] = {"wbs_code": wbs, "wbs_name": name,
                                "parent_wbs": parent_wbs, "level": level}
            if task.getSummary():
                continue

            start = _as_date(task.getStart())
            finish = _as_date(task.getFinish())
            baseline_start = _as_date(task.getBaselineStart()) or start
            baseline_finish = _as_date(task.getBaselineFinish()) or finish
            actual_start = _as_date(task.getActualStart())
            actual_finish = _as_date(task.getActualFinish())
            pct = _clamp(float(task.getPercentageComplete() or 0.0) / 100.0)
            critical = bool(task.getCritical())
            slack_days = _as_days(task.getTotalSlack())

            planned = pct
            if status_date is not None and baseline_start and baseline_finish:
                span = (baseline_finish - baseline_start).days
                if span > 0:
                    planned = _clamp((status_date - baseline_start).days / span)
                else:
                    planned = pct

            activity_id = f"ACT-{int(task.getID() or 0):04d}"
            schedule_rows.append([
                activity_id, wbs, name, "",  # activity_id..discipline
                _cell_date(baseline_start), _cell_date(baseline_finish),
                _cell_date(actual_start), _cell_date(actual_finish),
                planned, pct, round(slack_days), critical,
                _progress_status(pct),
            ])

            baseline_cost = float(task.getBaselineCost() or 0.0)
            actual_cost = float(task.getActualCost() or 0.0)
            budget_by_wbs[wbs] = budget_by_wbs.get(wbs, 0.0) + baseline_cost
            actual_by_wbs[wbs] = actual_by_wbs.get(wbs, 0.0) + actual_cost

        book = Workbook()
        wb = book.active
        wb.title = "Schedule"

        # ---- Project_Info (key/value pairs, header on row 1) ----------- #
        data_date = status_date or project_start or date.today()
        ws_info = book.create_sheet("Project_Info")
        ws_info.append(["project_field", "value"])
        ws_info.append(["project_id", ""])  # empty → falls back to project.code
        ws_info.append(["project_name", title])
        ws_info.append(["data_date", _cell_date(data_date)])
        ws_info.append(["dataset_version", "0.1"])

        def _sheet(name: str) -> Any:
            ws = book.create_sheet(name)
            ws.append([DEFAULT_SHEET_TITLE])
            ws.append([])
            return ws

        # ---- Schedule (header row 3, data from row 4) ------------------ #
        schedule_headers = [
            "activity_id", "wbs_code", "activity_name", "discipline",
            "baseline_start", "baseline_finish", "actual_start", "actual_finish",
            "planned_progress", "actual_progress", "total_float_days",
            "critical", "status",
        ]
        for row in ([DEFAULT_SHEET_TITLE], [], schedule_headers, *schedule_rows):
            wb.append(row)

        # ---- WBS ------------------------------------------------------- #
        ws_wbs = _sheet("WBS")
        wbs_headers = ["wbs_code", "wbs_name", "parent_wbs", "discipline", "level"]
        ws_wbs.append(wbs_headers)
        for entry in sorted(wbs_map.values(), key=lambda e: e["wbs_code"]):
            ws_wbs.append([entry["wbs_code"], entry["wbs_name"],
                           entry["parent_wbs"], "", entry["level"]])

        # ---- Progress (one snapshot row per WBS at the status date) ---- #
        ws_progress = _sheet("Progress")
        progress_headers = [
            "progress_id", "period", "wbs_code", "planned_progress",
            "actual_progress", "variance", "status",
        ]
        ws_progress.append(progress_headers)
        if status_date is not None:
            leaf_by_wbs: dict[str, list[list[Any]]] = {}
            for row in schedule_rows:
                leaf_by_wbs.setdefault(row[1], []).append(row)
            for wbs, rows in sorted(leaf_by_wbs.items()):
                planned = sum(r[8] for r in rows) / len(rows)
                actual = sum(r[9] for r in rows) / len(rows)
                ws_progress.append([
                    f"PRG-{wbs}", _cell_date(status_date), wbs,
                    round(planned, 4), round(actual, 4),
                    round(actual - planned, 4), _PROGRESS_STATUS,
                ])

        # ---- Budget (baseline cost aggregated per WBS) ----------------- #
        ws_budget = _sheet("Budget")
        budget_headers = [
            "budget_id", "wbs_code", "cost_code", "description",
            "budget_amount", "status", "effective_date",
        ]
        ws_budget.append(budget_headers)
        for wbs, amount in sorted(budget_by_wbs.items()):
            if amount <= 0:
                continue
            ws_budget.append([
                f"BDG-{wbs}", wbs, "", wbs_map.get(wbs, {}).get("wbs_name", ""),
                round(amount, 2), _PROGRESS_STATUS,
                _cell_date(project_start or date.today()),
            ])

        # ---- Actual_Cost (actual cost aggregated per WBS) -------------- #
        ws_actual = _sheet("Actual_Cost")
        actual_headers = [
            "transaction_id", "transaction_date", "wbs_code", "cost_code",
            "vendor_id", "vendor_name", "po_number", "description",
            "actual_amount", "status",
        ]
        ws_actual.append(actual_headers)
        for wbs, amount in sorted(actual_by_wbs.items()):
            if amount <= 0:
                continue
            ws_actual.append([
                f"AC-{wbs}", _cell_date(status_date or project_start or date.today()),
                wbs, "", "", "", "", wbs_map.get(wbs, {}).get("wbs_name", ""),
                round(amount, 2), _PROGRESS_STATUS,
            ])

        # ---- Commitments (empty governed sheet — MS Project has no PO data) ---- #
        ws_comm = _sheet("Commitments")
        commitment_headers = [
            "commitment_id", "wbs_code", "po_number", "vendor_id",
            "vendor_name", "committed_amount", "invoiced_amount",
            "status", "commitment_date",
        ]
        ws_comm.append(commitment_headers)

        buffer = BytesIO()
        book.save(buffer)
        return buffer.getvalue()


def build_mpp_converter() -> MppConverter | None:
    """Instantiate the converter only when the runtime can actually use it.

    Returns ``None`` on serverless runtimes (no JVM / no mpxj package) so the
    API can reject ``.mpp`` uploads with a clear 'worker required' error.
    """
    import importlib.util
    import os
    import shutil

    java_home = os.environ.get("CONTROLCHECK_JAVA_HOME", "").strip()
    java_bin = None
    if java_home:
        candidate = Path(java_home) / "bin" / "java"
        if candidate.is_file():
            java_bin = str(candidate)
    if java_bin is None:
        java_bin = shutil.which("java")
    if java_bin is None:
        return None
    if importlib.util.find_spec("mpxj") is None or importlib.util.find_spec("jpype") is None:
        return None
    return MppConverter(java_home=os.path.dirname(os.path.dirname(java_bin)))
