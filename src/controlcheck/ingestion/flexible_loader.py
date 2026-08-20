"""Smart Flexible Excel Workbook Loader & Column Auto-Mapper.

Enables universal ingestion of custom Excel sheets, SAP ERP exports,
Primavera P6 exports, and Indonesian column conventions into canonical ProjectDataset.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import re
from typing import Any, BinaryIO

import openpyxl

from ..logging import get_logger
from ..models import (
    ActualCostRecord,
    BudgetRecord,
    CommitmentRecord,
    ProgressRecord,
    ProjectDataset,
    ProjectInfo,
    ScheduleActivity,
    SourceRef,
    WBSNode,
)

logger = get_logger("flexible_loader")


def _norm(name: str) -> str:
    """Normalize string by removing whitespace, punctuation, and lowercasing."""
    return re.sub(r"[^a-z0-9]", "", str(name).lower())


# Canonical Sheet Name Aliases
SHEET_ALIASES: dict[str, list[str]] = {
    "Project_Info": [
        "projectinfo", "projectmetadata", "metadata", "project", "proyek", "info", "general", "cover"
    ],
    "WBS": [
        "wbs", "workbreakdownstructure", "strukturwbs", "wbslist", "daftarwbs", "wbsstructure"
    ],
    "Budget": [
        "budget", "anggaran", "rab", "costbudget", "budgetplan", "anggaranbiaya", "budgetdata"
    ],
    "Actual_Cost": [
        "actualcost", "actualcosts", "actuals", "realisasi", "realisasibiaya", "pengeluaran",
        "biayaaktual", "costactuals", "transaksibiaya", "actual"
    ],
    "Commitments": [
        "commitments", "commitment", "po", "purchaseorders", "komitmen", "kontrakvendor", "purchaseorder"
    ],
    "Schedule": [
        "schedule", "jadwal", "activities", "gantt", "primavera", "p6export", "tasklist",
        "timeline", "jadwalproyek", "tasks"
    ],
    "Progress": [
        "progress", "progres", "kurvas", "scurve", "physicalprogress", "kemajuanfisik",
        "progresfisik", "kurva"
    ],
}

# Canonical Column Synonym Dictionaries
COLUMN_SYNONYMS: dict[str, list[str]] = {
    # WBS
    "wbs_code": ["wbscode", "wbs", "wbsid", "kodewbs", "wbselement", "wbsno", "wbsnumber", "kode"],
    "wbs_name": ["wbsname", "namawbs", "wbsdescription", "deskripsiwbs", "uraianpekerjaan", "name", "nama", "description", "deskripsi"],
    "parent_wbs": ["parentwbs", "parent", "indukwbs", "wbsinduk", "parentcode", "parentid"],
    "discipline": ["discipline", "disiplin", "bidang", "kategori", "category", "workpackage"],
    "level": ["level", "tingkat", "hierarki", "hierarchy", "wbslevel"],

    # Budget
    "budget_id": ["budgetid", "idbudget", "idanggaran", "budgetcode", "nobudget", "kodebudget"],
    "cost_code": ["costcode", "kodebiaya", "kodeakun", "glaccount", "accountcode", "costelement", "akun"],
    "budget_amount": [
        "budgetamount", "budget", "anggaran", "nilairab", "currentbudget", "approvedbudget",
        "totalbudget", "nilaipagu", "nilaianggaran", "amount", "nilai"
    ],
    "status": ["status", "keadaan", "state", "keterangan"],
    "effective_date": ["effectivedate", "date", "tanggal", "tglberlaku", "postingdate", "tgllapor"],

    # Actual Cost
    "transaction_id": ["transactionid", "idtransaksi", "notransaksi", "transid", "voucher", "docnumber", "refno", "id"],
    "transaction_date": ["transactiondate", "postingdate", "tanggal", "tgltransaksi", "date", "docdate", "tglposting"],
    "vendor_id": ["vendorid", "idvendor", "supplierid", "kodevendor", "vendorcode", "rekanan"],
    "vendor_name": ["vendorname", "namavendor", "suppliername", "namarekanan", "vendor", "supplier", "rekanan"],
    "po_number": ["ponumber", "nopo", "purchasingorder", "nospk", "nokontrak", "poid", "contractno", "nopo"],
    "actual_amount": [
        "actualamount", "actualcost", "realisasi", "biayaaktual", "nilairealisasi", "amount",
        "totalcost", "pengeluaran", "nilai"
    ],

    # Commitments
    "commitment_id": ["commitmentid", "idkomitmen", "idpo", "nopo", "poid", "id"],
    "committed_amount": ["committedamount", "committedcost", "nilaipo", "nilaikomitmen", "contractamount", "amount", "nilai"],
    "invoiced_amount": ["invoicedamount", "invoicedcost", "nilaiinvoice", "tagihan", "billedamount", "invoiced"],
    "commitment_date": ["commitmentdate", "podate", "tglpo", "tglkontrak", "date", "tanggal"],

    # Schedule
    "activity_id": ["activityid", "actid", "taskid", "idaktivitas", "kodeaktivitas", "activitycode", "id"],
    "activity_name": ["activityname", "actname", "taskname", "namaaktivitas", "deskripsiaktivitas", "uraian", "task", "activity", "nama"],
    "baseline_start": ["baselinestart", "blstart", "targetstart", "tglmulairencana", "plannedstart", "startdate", "mulai", "tglmulai"],
    "baseline_finish": ["baselinefinish", "blfinish", "targetfinish", "tglselesairencana", "plannedfinish", "finishdate", "selesai", "tglselesai"],
    "actual_start": ["actualstart", "actstart", "tglmulaiaktual", "actualstartdate", "mulaiactual"],
    "actual_finish": ["actualfinish", "actfinish", "tglselesaiaktual", "actualfinishdate", "selesaiactual"],
    "planned_progress": ["plannedprogress", "plannedprogresspct", "planned", "rencana", "bobotrencana", "targetprogress", "plannedpct", "rencana"],
    "actual_progress": ["actualprogress", "actualprogresspct", "actual", "realisasi", "bobotrealisasi", "physicalprogress", "actualpct", "realisasi"],
    "total_float_days": ["totalfloatdays", "totalfloat", "tf", "float", "slack", "totalslack", "sisafloat"],
    "critical": ["critical", "iscritical", "jalurkritis", "kritis"],

    # Progress
    "progress_id": ["progressid", "idprogres", "idkemajuan", "id"],
    "period": ["period", "cutoffdate", "tanggal", "periode", "tglcutoff", "tgl"],
    "variance": ["variance", "varians", "deviasi", "gap"],
}


def _match_sheet_name(available_sheets: list[str], canonical_target: str) -> str | None:
    """Find sheet matching target canonical name or synonyms with substring/fuzzy match."""
    target_clean = _norm(canonical_target)
    aliases = SHEET_ALIASES.get(canonical_target, [target_clean])

    # 1. Exact match
    for sheet in available_sheets:
        norm_s = _norm(sheet)
        if norm_s == target_clean or norm_s in aliases:
            return sheet

    # 2. Fuzzy substring match
    for sheet in available_sheets:
        norm_s = _norm(sheet)
        for alias in aliases:
            if len(alias) >= 3 and (alias in norm_s or norm_s in alias):
                return sheet
    return None


def _map_row_to_canonical(row_dict: dict[str, Any], canonical_fields: list[str]) -> dict[str, Any]:
    """Map raw row dictionary keys to canonical field names using synonym dictionary."""
    mapped: dict[str, Any] = {}
    norm_lookup = {_norm(k): v for k, v in row_dict.items() if k is not None}

    for target_field in canonical_fields:
        target_norm = _norm(target_field)
        synonyms = COLUMN_SYNONYMS.get(target_field, [target_norm])
        matched_val = None

        # 1. Exact synonym match
        for syn in synonyms:
            if syn in norm_lookup:
                matched_val = norm_lookup[syn]
                break

        # 2. Substring synonym match
        if matched_val is None:
            for k_norm, val in norm_lookup.items():
                for syn in synonyms:
                    if len(syn) >= 3 and (syn in k_norm or k_norm in syn):
                        matched_val = val
                        break
                if matched_val is not None:
                    break

        mapped[target_field] = matched_val
    return mapped



def _parse_date(value: Any, fallback: date | None = None) -> date:
    if value is None:
        return fallback or date.today()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    val_str = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(val_str).date()
    except ValueError:
        return fallback or date.today()


def _parse_decimal(value: Any, default: Decimal = Decimal(0)) -> Decimal:
    if value is None or value == "":
        return default
    try:
        # Strip potential currency symbols, commas, spaces
        cleaned = re.sub(r"[^\d.-]", "", str(value))
        return Decimal(cleaned) if cleaned else default
    except Exception:
        return default


def _parse_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        cleaned = re.sub(r"[^\d.-]", "", str(value))
        val = float(cleaned) if cleaned else default
        # If progress is between 0 and 1 (e.g. 0.85), convert to percentage (85.0) if applicable
        return val
    except Exception:
        return default


def _read_sheet_rows(sheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, list[str]]:
    """Detect header row and column names dynamically."""
    best_row = 1
    best_headers: list[str] = []
    for row_num, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
        non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if len(non_empty) > len(best_headers):
            best_headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(row)]
            best_row = row_num
    return best_row, best_headers


def load_flexible_workbook(source: Path | str | BinaryIO) -> ProjectDataset:
    """Load arbitrary Excel workbook with smart sheet & header mapping."""
    logger.info("Engaging Smart Flexible Workbook Auto-Mapper...")
    book = openpyxl.load_workbook(source, data_only=True, read_only=False)
    sheetnames = book.sheetnames

    # 1. Project Info / Metadata
    meta_sheet_name = _match_sheet_name(sheetnames, "Project_Info")
    project_id = "PRJ-FLEX-001"
    project_name = "Project Control Audit"
    data_date = date.today()
    dataset_version = "0.2"

    if meta_sheet_name:
        meta_sheet = book[meta_sheet_name]
        for row in meta_sheet.iter_rows(values_only=True):
            if len(row) >= 2 and row[0] is not None:
                k = _norm(row[0])
                v = row[1]
                if "projectid" in k and v:
                    project_id = str(v).strip()
                elif "projectname" in k and v:
                    project_name = str(v).strip()
                elif "datadate" in k and v:
                    data_date = _parse_date(v)
                elif "datasetversion" in k and v:
                    dataset_version = str(v).strip()

    # 2. WBS
    wbs_sheet_name = _match_sheet_name(sheetnames, "WBS")
    wbs_nodes: list[WBSNode] = []
    if wbs_sheet_name:
        sheet = book[wbs_sheet_name]
        h_row, headers = _read_sheet_rows(sheet)
        for r_num, row in enumerate(sheet.iter_rows(min_row=h_row + 1, values_only=True), start=h_row + 1):
            if not any(c is not None for c in row):
                continue
            row_dict = dict(zip(headers, row))
            mapped = _map_row_to_canonical(row_dict, ["wbs_code", "wbs_name", "parent_wbs", "discipline", "level"])
            wbs_code = str(mapped["wbs_code"] or f"WBS-{r_num}").strip()
            level_raw = mapped["level"]
            try:
                level = int(str(level_raw).strip()) if level_raw is not None else max(1, wbs_code.count(".") + 1)
            except Exception:
                level = max(1, wbs_code.count(".") + 1)

            wbs_nodes.append(WBSNode(
                wbs_code=wbs_code,
                wbs_name=str(mapped["wbs_name"] or wbs_code).strip(),
                parent_wbs=str(mapped["parent_wbs"]).strip() if mapped["parent_wbs"] else None,
                discipline=str(mapped["discipline"] or "GENERAL").strip(),
                level=level,
                source=SourceRef(sheet=wbs_sheet_name, row_number=r_num),
            ))


    # 3. Budget
    budget_sheet_name = _match_sheet_name(sheetnames, "Budget")
    budgets: list[BudgetRecord] = []
    if budget_sheet_name:
        sheet = book[budget_sheet_name]
        h_row, headers = _read_sheet_rows(sheet)
        for r_num, row in enumerate(sheet.iter_rows(min_row=h_row + 1, values_only=True), start=h_row + 1):
            if not any(c is not None for c in row):
                continue
            row_dict = dict(zip(headers, row))
            mapped = _map_row_to_canonical(row_dict, ["budget_id", "wbs_code", "cost_code", "description", "budget_amount", "status", "effective_date"])
            budgets.append(BudgetRecord(
                budget_id=str(mapped["budget_id"] or f"BDG-{r_num}").strip(),
                wbs_code=str(mapped["wbs_code"] or "1.0").strip(),
                cost_code=str(mapped["cost_code"] or "LABOR").strip(),
                description=str(mapped["description"] or "Budget Allocation").strip(),
                budget_amount=_parse_decimal(mapped["budget_amount"]),
                status=str(mapped["status"] or "APPROVED").strip(),
                effective_date=_parse_date(mapped["effective_date"], data_date),
                source=SourceRef(sheet=budget_sheet_name, row_number=r_num),
            ))

    # 4. Actual Costs
    actual_sheet_name = _match_sheet_name(sheetnames, "Actual_Cost")
    actual_costs: list[ActualCostRecord] = []
    if actual_sheet_name:
        sheet = book[actual_sheet_name]
        h_row, headers = _read_sheet_rows(sheet)
        for r_num, row in enumerate(sheet.iter_rows(min_row=h_row + 1, values_only=True), start=h_row + 1):
            if not any(c is not None for c in row):
                continue
            row_dict = dict(zip(headers, row))
            mapped = _map_row_to_canonical(row_dict, ["transaction_id", "transaction_date", "wbs_code", "cost_code", "vendor_id", "vendor_name", "po_number", "description", "actual_amount", "status"])
            actual_costs.append(ActualCostRecord(
                transaction_id=str(mapped["transaction_id"] or f"TX-{r_num}").strip(),
                transaction_date=_parse_date(mapped["transaction_date"], data_date),
                wbs_code=str(mapped["wbs_code"] or "1.0").strip(),
                cost_code=str(mapped["cost_code"] or "COST").strip(),
                vendor_id=str(mapped["vendor_id"]).strip() if mapped["vendor_id"] else None,
                vendor_name=str(mapped["vendor_name"]).strip() if mapped["vendor_name"] else None,
                po_number=str(mapped["po_number"]).strip() if mapped["po_number"] else None,
                description=str(mapped["description"] or "Actual Incurred Cost").strip(),
                actual_amount=_parse_decimal(mapped["actual_amount"]),
                status=str(mapped["status"] or "POSTED").strip(),
                source=SourceRef(sheet=actual_sheet_name, row_number=r_num),
            ))

    # 5. Commitments
    comm_sheet_name = _match_sheet_name(sheetnames, "Commitments")
    commitments: list[CommitmentRecord] = []
    if comm_sheet_name:
        sheet = book[comm_sheet_name]
        h_row, headers = _read_sheet_rows(sheet)
        for r_num, row in enumerate(sheet.iter_rows(min_row=h_row + 1, values_only=True), start=h_row + 1):
            if not any(c is not None for c in row):
                continue
            row_dict = dict(zip(headers, row))
            mapped = _map_row_to_canonical(row_dict, ["commitment_id", "wbs_code", "po_number", "vendor_id", "vendor_name", "committed_amount", "invoiced_amount", "status", "commitment_date"])
            commitments.append(CommitmentRecord(
                commitment_id=str(mapped["commitment_id"] or f"COM-{r_num}").strip(),
                wbs_code=str(mapped["wbs_code"] or "1.0").strip(),
                po_number=str(mapped["po_number"] or f"PO-{r_num}").strip(),
                vendor_id=str(mapped["vendor_id"] or "V001").strip(),
                vendor_name=str(mapped["vendor_name"] or "Vendor Partner").strip(),
                committed_amount=_parse_decimal(mapped["committed_amount"]),
                invoiced_amount=_parse_decimal(mapped["invoiced_amount"]),
                status=str(mapped["status"] or "ACTIVE").strip(),
                commitment_date=_parse_date(mapped["commitment_date"], data_date),
                source=SourceRef(sheet=comm_sheet_name, row_number=r_num),
            ))

    # 6. Schedule
    sched_sheet_name = _match_sheet_name(sheetnames, "Schedule")
    schedule: list[ScheduleActivity] = []
    if sched_sheet_name:
        sheet = book[sched_sheet_name]
        h_row, headers = _read_sheet_rows(sheet)
        for r_num, row in enumerate(sheet.iter_rows(min_row=h_row + 1, values_only=True), start=h_row + 1):
            if not any(c is not None for c in row):
                continue
            row_dict = dict(zip(headers, row))
            mapped = _map_row_to_canonical(row_dict, [
                "activity_id", "wbs_code", "activity_name", "discipline", "baseline_start",
                "baseline_finish", "actual_start", "actual_finish", "planned_progress",
                "actual_progress", "total_float_days", "critical", "status"
            ])
            crit_val = mapped["critical"]
            is_crit = str(crit_val).lower().strip() in ("true", "1", "yes", "y", "critical", "kritis", "ya") if crit_val is not None else False

            schedule.append(ScheduleActivity(
                activity_id=str(mapped["activity_id"] or f"ACT-{r_num}").strip(),
                wbs_code=str(mapped["wbs_code"] or "1.0").strip(),
                activity_name=str(mapped["activity_name"] or f"Activity {r_num}").strip(),
                discipline=str(mapped["discipline"] or "GENERAL").strip(),
                baseline_start=_parse_date(mapped["baseline_start"], data_date),
                baseline_finish=_parse_date(mapped["baseline_finish"], data_date),
                actual_start=_parse_date(mapped["actual_start"]) if mapped["actual_start"] else None,
                actual_finish=_parse_date(mapped["actual_finish"]) if mapped["actual_finish"] else None,
                planned_progress=_parse_float(mapped["planned_progress"]),
                actual_progress=_parse_float(mapped["actual_progress"]),
                total_float_days=_parse_float(mapped["total_float_days"]),
                critical=is_crit,
                status=str(mapped["status"] or "IN_PROGRESS").strip(),
                source=SourceRef(sheet=sched_sheet_name, row_number=r_num),
            ))

    # 7. Progress
    prog_sheet_name = _match_sheet_name(sheetnames, "Progress")
    progress: list[ProgressRecord] = []
    if prog_sheet_name:
        sheet = book[prog_sheet_name]
        h_row, headers = _read_sheet_rows(sheet)
        for r_num, row in enumerate(sheet.iter_rows(min_row=h_row + 1, values_only=True), start=h_row + 1):
            if not any(c is not None for c in row):
                continue
            row_dict = dict(zip(headers, row))
            mapped = _map_row_to_canonical(row_dict, ["progress_id", "period", "wbs_code", "planned_progress", "actual_progress", "variance", "status"])
            plan_p = _parse_float(mapped["planned_progress"])
            act_p = _parse_float(mapped["actual_progress"])
            progress.append(ProgressRecord(
                progress_id=str(mapped["progress_id"] or f"PRG-{r_num}").strip(),
                period=_parse_date(mapped["period"], data_date),
                wbs_code=str(mapped["wbs_code"] or "1.0").strip(),
                planned_progress=plan_p,
                actual_progress=act_p,
                variance=_parse_float(mapped["variance"], act_p - plan_p),
                status=str(mapped["status"] or "SUBMITTED").strip(),
                source=SourceRef(sheet=prog_sheet_name, row_number=r_num),
            ))

    book.close()

    logger.info(
        "Smart Flexible Ingestion successfully mapped: %d WBS, %d Budgets, %d Actuals, %d Commitments, %d Schedule, %d Progress",
        len(wbs_nodes), len(budgets), len(actual_costs), len(commitments), len(schedule), len(progress)
    )

    return ProjectDataset(
        project=ProjectInfo(project_id=project_id, project_name=project_name),
        data_date=data_date,
        wbs_nodes=wbs_nodes,
        budgets=budgets,
        actual_costs=actual_costs,
        commitments=commitments,
        schedule=schedule,
        progress=progress,
        dataset_version=dataset_version,
    )
