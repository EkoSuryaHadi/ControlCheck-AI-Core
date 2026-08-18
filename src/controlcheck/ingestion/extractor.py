from __future__ import annotations

from base64 import b64encode
from datetime import date, datetime, time
from decimal import Decimal
from hashlib import sha256
from io import BytesIO
import json
import math
from typing import Any

import openpyxl

from .profile import MappingProfileV1
from .types import ExtractedRow, ExtractedWorkbook, TemplateIssue


def _json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, bytes):
        return b64encode(value).decode("ascii")
    return str(value)


def _header(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    return result or None


def _unique_value_keys(headers: tuple[Any, ...]) -> list[str]:
    counts: dict[str, int] = {}
    keys: list[str] = []
    for position, value in enumerate(headers, start=1):
        header = _header(value) or f"__unnamed_column_{position}"
        counts[header] = counts.get(header, 0) + 1
        occurrence = counts[header]
        keys.append(header if occurrence == 1 else f"{header}__duplicate_{occurrence}")
    return keys


def _source_key(domain: str, source_row_number: int, values: dict[str, Any]) -> str:
    payload = json.dumps(
        {"domain": domain, "source_row_number": source_row_number, "values": values},
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
    )
    return sha256(payload.encode("utf-8")).hexdigest()


def _project_values(book: openpyxl.Workbook) -> dict[str, Any]:
    if "Project_Info" not in book.sheetnames:
        return {}

    result: dict[str, Any] = {}
    for row in book["Project_Info"].iter_rows(values_only=True):
        if not row:
            continue
        key = _header(row[0])
        if key is not None:
            result[key] = _json_safe(row[1] if len(row) > 1 else None)
    return result


def _header_issues(domain: str, sheet_name: str, headers: tuple[Any, ...], profile) -> list[TemplateIssue]:
    normalized_headers = [_header(value) for value in headers]
    present_headers = {header for header in normalized_headers if header is not None}
    expected_headers = {
        column.source_header for column in profile.columns.values()
    }
    required_headers = {
        column.source_header
        for column in profile.columns.values()
        if column.required
    }
    duplicates = sorted(
        {
            header
            for header in present_headers
            if normalized_headers.count(header) > 1
        }
    )
    missing = sorted(required_headers - present_headers)
    unexpected = sorted(present_headers - expected_headers)
    issues: list[TemplateIssue] = []
    if duplicates:
        issues.append(
            TemplateIssue(
                code="duplicate_header",
                message=f"{sheet_name} has duplicate headers: {', '.join(duplicates)}",
                domain=domain,
                sheet_name=sheet_name,
            )
        )
    if missing:
        issues.append(
            TemplateIssue(
                code="missing_required_column",
                message=f"{sheet_name} is missing required columns: {', '.join(missing)}",
                domain=domain,
                sheet_name=sheet_name,
            )
        )
    if unexpected:
        issues.append(
            TemplateIssue(
                code="unexpected_header",
                message=f"{sheet_name} has unexpected headers: {', '.join(unexpected)}",
                domain=domain,
                sheet_name=sheet_name,
            )
        )
    return issues


def extract_workbook(data: bytes, profile: MappingProfileV1) -> ExtractedWorkbook:
    """Extract governed workbook rows without validating or coercing their values."""
    book = openpyxl.load_workbook(BytesIO(data), data_only=True)
    try:
        rows_by_domain: dict[str, list[ExtractedRow]] = {}
        template_errors: list[TemplateIssue] = []
        for domain, domain_profile in profile.domains.items():
            sheet_name = domain_profile.sheet_name
            if sheet_name not in book.sheetnames:
                rows_by_domain[domain] = []
                template_errors.append(
                    TemplateIssue(
                        code="missing_sheet",
                        message=f"Missing governed sheet: {sheet_name}",
                        domain=domain,
                        sheet_name=sheet_name,
                    )
                )
                continue

            sheet = book[sheet_name]
            header_row = next(
                sheet.iter_rows(
                    min_row=domain_profile.header_row,
                    max_row=domain_profile.header_row,
                    values_only=True,
                )
            )
            template_errors.extend(
                _header_issues(domain, sheet_name, header_row, domain_profile)
            )
            keys = _unique_value_keys(header_row)
            rows: list[ExtractedRow] = []
            for source_row_number, row in enumerate(
                sheet.iter_rows(min_row=domain_profile.header_row + 1, values_only=True),
                start=domain_profile.header_row + 1,
            ):
                if not any(value is not None for value in row):
                    continue
                values = {
                    key: _json_safe(value)
                    for key, value in zip(keys, row)
                }
                rows.append(
                    ExtractedRow(
                        domain=domain,
                        sheet_name=sheet_name,
                        source_row_number=source_row_number,
                        values=values,
                        source_key=_source_key(domain, source_row_number, values),
                    )
                )
            rows_by_domain[domain] = rows

        template_errors.sort(
            key=lambda issue: (issue.domain, issue.sheet_name, issue.code, issue.message)
        )
        return ExtractedWorkbook(
            project_values=_project_values(book),
            rows_by_domain=rows_by_domain,
            template_errors=template_errors,
            workbook_sha256=sha256(data).hexdigest(),
        )
    finally:
        book.close()
