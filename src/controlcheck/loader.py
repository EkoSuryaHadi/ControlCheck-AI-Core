from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, BinaryIO, TypeVar

import openpyxl
from pydantic import BaseModel

from .models import (
    ActualCostRecord,
    BudgetRecord,
    CommitmentRecord,
    ProgressRecord,
    ProjectDataset,
    ProjectInfo,
    ScheduleActivity,
    SourceRef,
    WBSNode,
)


REQUIRED_COLUMNS = {
    "WBS": {"wbs_code", "wbs_name", "parent_wbs", "discipline", "level"},
    "Budget": {"budget_id", "wbs_code", "cost_code", "description", "budget_amount", "status", "effective_date"},
    "Actual_Cost": {"transaction_id", "transaction_date", "wbs_code", "cost_code", "vendor_id", "vendor_name", "po_number", "description", "actual_amount", "status"},
    "Commitments": {"commitment_id", "wbs_code", "po_number", "vendor_id", "vendor_name", "committed_amount", "invoiced_amount", "status", "commitment_date"},
    "Schedule": {"activity_id", "wbs_code", "activity_name", "discipline", "baseline_start", "baseline_finish", "actual_start", "actual_finish", "planned_progress", "actual_progress", "total_float_days", "critical", "status"},
    "Progress": {"progress_id", "period", "wbs_code", "planned_progress", "actual_progress", "variance", "status"},
}


class WorkbookSchemaError(ValueError):
    def __init__(self, code: str, message: str):
        self.code = code
        super().__init__(message)


T = TypeVar("T", bound=BaseModel)


def _text(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    return result or None


def _date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.fromisoformat(str(value)).date()


def _records(sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, Any]]:
    header_row = None
    headers: list[str] = []
    required = REQUIRED_COLUMNS[sheet.title]
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        candidate = [str(value).strip() if value is not None else "" for value in row]
        if required <= set(candidate):
            header_row, headers = row_number, candidate
            break
    if header_row is None:
        missing = ", ".join(sorted(required))
        raise WorkbookSchemaError("missing_columns", f"{sheet.title} is missing required columns: {missing}")
    missing = required - set(headers)
    if missing:
        raise WorkbookSchemaError("missing_columns", f"{sheet.title} is missing required columns: {', '.join(sorted(missing))}")
    result = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(value is not None for value in row):
            continue
        record = dict(zip(headers, row))
        record["source"] = SourceRef(sheet=sheet.title, row_number=row_number)
        result.append(record)
    return result


def _project_info(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, Any]:
    values = {}
    for key, value, *_ in sheet.iter_rows(min_row=2, values_only=True):
        if key is not None:
            values[str(key).strip()] = value
    return values


def load_workbook(path: Path | str | BinaryIO) -> ProjectDataset:
    source = path
    if isinstance(path, (str, Path)):
        source = Path(path)
        if not source.exists():
            raise WorkbookSchemaError("file_not_found", f"Workbook not found: {source}")
    # Non-streaming mode releases the Windows file handle reliably after close,
    # which is required by the API's bounded temporary-file lifecycle.
    book = openpyxl.load_workbook(source, data_only=True, read_only=False)
    missing_sheets = set(REQUIRED_COLUMNS) - set(book.sheetnames)
    if missing_sheets:
        raise WorkbookSchemaError("missing_sheets", f"Missing required sheets: {', '.join(sorted(missing_sheets))}")
    info = _project_info(book["Project_Info"])

    wbs = [WBSNode(**{**r, "parent_wbs": _text(r["parent_wbs"]), "discipline": _text(r["discipline"])}) for r in _records(book["WBS"])]
    budgets = [BudgetRecord(**{**r, "wbs_code": _text(r["wbs_code"]), "cost_code": _text(r["cost_code"]), "budget_amount": Decimal(str(r["budget_amount"])), "effective_date": _date(r["effective_date"])}) for r in _records(book["Budget"])]
    actuals = [ActualCostRecord(**{**r, "transaction_date": _date(r["transaction_date"]), "wbs_code": _text(r["wbs_code"]), "cost_code": _text(r["cost_code"]), "vendor_id": _text(r["vendor_id"]), "vendor_name": _text(r["vendor_name"]), "po_number": _text(r["po_number"]), "actual_amount": Decimal(str(r["actual_amount"]))}) for r in _records(book["Actual_Cost"])]
    commitments = [CommitmentRecord(**{**r, "wbs_code": _text(r["wbs_code"]), "po_number": _text(r["po_number"]), "vendor_id": _text(r["vendor_id"]), "vendor_name": _text(r["vendor_name"]), "committed_amount": Decimal(str(r["committed_amount"])), "invoiced_amount": Decimal(str(r["invoiced_amount"])), "commitment_date": _date(r["commitment_date"])}) for r in _records(book["Commitments"])]
    schedule = [ScheduleActivity(**{**r, "wbs_code": _text(r["wbs_code"]), "discipline": _text(r["discipline"]), "baseline_start": _date(r["baseline_start"]), "baseline_finish": _date(r["baseline_finish"]), "actual_start": _date(r["actual_start"]) if r["actual_start"] else None, "actual_finish": _date(r["actual_finish"]) if r["actual_finish"] else None}) for r in _records(book["Schedule"])]
    progress = [ProgressRecord(**{**r, "period": _date(r["period"]), "wbs_code": _text(r["wbs_code"])}) for r in _records(book["Progress"])]

    result = ProjectDataset(
        project=ProjectInfo(project_id=str(info["project_id"]), project_name=str(info["project_name"])),
        data_date=_date(info["data_date"]),
        wbs_nodes=wbs,
        budgets=budgets,
        actual_costs=actuals,
        commitments=commitments,
        schedule=schedule,
        progress=progress,
        dataset_version=str(info.get("dataset_version", "0.1")),
    )
    book.close()
    return result
